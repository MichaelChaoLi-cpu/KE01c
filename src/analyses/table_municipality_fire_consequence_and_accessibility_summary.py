#!/usr/bin/env python3
"""Summarize conditional fire consequence and accessibility by municipality.

The table aggregates the accepted 125 m consequence and event-specific
accessibility outputs. Grid cells that cross administrative boundaries are
assigned to the unit with the largest overlapping area; non-intersecting
coastal slivers are assigned to the nearest administrative polygon and are
reported in the diagnostics.
"""

from __future__ import annotations

from pathlib import Path

import geopandas as gpd
import numpy as np
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd


ROOT = Path(__file__).resolve().parents[2]
CONSEQUENCE_PATH = ROOT / "data/results/derived/fire_consequence_125m.parquet"
ADMIN_PATH = ROOT / "data/processed/administrative_areas_preprocessed.parquet"
OUTPUT = ROOT / "data/results/tables/Table_municipality_fire_consequence_and_accessibility_summary.xlsx"

PROJECTED_CRS = 6670
HIGH_CONSEQUENCE_THRESHOLD = 0.90
TIMELY_RESPONSE_THRESHOLD_MIN = 10.0

ENGLISH_MUNICIPALITY = {
    "43101": "Kumamoto City, Chuo Ward",
    "43102": "Kumamoto City, Higashi Ward",
    "43103": "Kumamoto City, Nishi Ward",
    "43104": "Kumamoto City, Minami Ward",
    "43105": "Kumamoto City, Kita Ward",
    "43202": "Yatsushiro City",
    "43203": "Hitoyoshi City",
    "43204": "Arao City",
    "43205": "Minamata City",
    "43206": "Tamana City",
    "43208": "Yamaga City",
    "43210": "Kikuchi City",
    "43211": "Uto City",
    "43212": "Kamiamakusa City",
    "43213": "Uki City",
    "43214": "Aso City",
    "43215": "Amakusa City",
    "43216": "Koshi City",
    "43348": "Misato Town",
    "43364": "Gyokuto Town",
    "43367": "Nankan Town",
    "43368": "Nagasu Town",
    "43369": "Nagomi Town",
    "43403": "Ozu Town",
    "43404": "Kikuyo Town",
    "43423": "Minamioguni Town",
    "43424": "Oguni Town",
    "43425": "Ubuyama Village",
    "43428": "Takamori Town",
    "43432": "Nishihara Village",
    "43433": "Minamiaso Village",
    "43441": "Mifune Town",
    "43442": "Kashima Town",
    "43443": "Mashiki Town",
    "43444": "Kosa Town",
    "43447": "Yamato Town",
    "43468": "Hikawa Town",
    "43482": "Ashikita Town",
    "43484": "Tsunagi Town",
    "43501": "Nishiki Town",
    "43505": "Taragi Town",
    "43506": "Yunomae Town",
    "43507": "Mizukami Village",
    "43510": "Sagara Village",
    "43511": "Itsuki Village",
    "43512": "Yamae Village",
    "43513": "Kuma Village",
    "43514": "Asagiri Town",
    "43531": "Reihoku Town",
}


def weighted_mean(values: pd.Series, weights: pd.Series) -> float:
    """Return a finite non-negative weighted mean or NaN."""
    values = pd.to_numeric(values, errors="coerce")
    weights = pd.to_numeric(weights, errors="coerce")
    valid = values.notna() & weights.notna() & (weights >= 0)
    if not valid.any() or float(weights.loc[valid].sum()) <= 0:
        return float("nan")
    return float(np.average(values.loc[valid], weights=weights.loc[valid]))


def assign_municipalities(
    cells: gpd.GeoDataFrame,
    administrative: gpd.GeoDataFrame,
) -> tuple[pd.DataFrame, dict[str, float]]:
    """Assign cells by maximum overlap, with nearest fallback for coastal slivers."""
    left = cells[["Mesh Code", "Geometry"]].copy()
    right = administrative[
        ["Municipality Code", "Municipality Label", "Geometry"]
    ].copy()
    points = left.copy()
    points["Geometry"] = points.geometry.representative_point()
    assigned = gpd.sjoin(points, right, how="left", predicate="within")
    assigned = (
        assigned.sort_values(["Mesh Code", "Municipality Code"], kind="stable")
        .drop_duplicates("Mesh Code")
        [["Mesh Code", "Municipality Code", "Municipality Label"]]
        .copy()
    )

    boundary_codes = assigned.loc[
        assigned["Municipality Code"].isna(), "Mesh Code"
    ].copy()
    if len(boundary_codes):
        boundary_cells = left.loc[left["Mesh Code"].isin(boundary_codes)].copy()
        candidates = gpd.sjoin(
            boundary_cells,
            right,
            how="left",
            predicate="intersects",
        ).reset_index(drop=True)
        candidates["Overlap Area (m2)"] = -1.0
        valid = candidates["index_right"].notna()
        candidate_geometry = gpd.GeoSeries(
            candidates.loc[valid, "index_right"].map(right.geometry),
            index=candidates.index[valid],
            crs=left.crs,
        )
        candidates.loc[valid, "Overlap Area (m2)"] = (
            candidates.loc[valid].geometry.intersection(candidate_geometry).area
        )
        overlap_assignment = (
            candidates.loc[valid]
            .sort_values(
                ["Mesh Code", "Overlap Area (m2)", "Municipality Code"],
                ascending=[True, False, True],
                kind="stable",
            )
            .drop_duplicates("Mesh Code")
            [["Mesh Code", "Municipality Code", "Municipality Label"]]
            .set_index("Mesh Code")
        )
        assigned = assigned.set_index("Mesh Code")
        assigned.loc[
            overlap_assignment.index,
            ["Municipality Code", "Municipality Label"],
        ] = overlap_assignment[["Municipality Code", "Municipality Label"]]
        assigned = assigned.reset_index()

    unmatched_codes = assigned.loc[assigned["Municipality Code"].isna(), "Mesh Code"]
    maximum_fallback_distance = 0.0
    if len(unmatched_codes):
        unmatched = left.loc[left["Mesh Code"].isin(unmatched_codes)].copy()
        unmatched["Geometry"] = unmatched.geometry.representative_point()
        nearest = gpd.sjoin_nearest(
            unmatched,
            right,
            how="left",
            distance_col="Administrative Distance (m)",
        )
        nearest = (
            nearest.sort_values(
                ["Mesh Code", "Administrative Distance (m)", "Municipality Code"],
                kind="stable",
            )
            .drop_duplicates("Mesh Code")
            [[
                "Mesh Code",
                "Municipality Code",
                "Municipality Label",
                "Administrative Distance (m)",
            ]]
        )
        maximum_fallback_distance = float(nearest["Administrative Distance (m)"].max())
        assigned = assigned.set_index("Mesh Code")
        nearest = nearest.set_index("Mesh Code")
        assigned.loc[nearest.index, ["Municipality Code", "Municipality Label"]] = nearest[
            ["Municipality Code", "Municipality Label"]
        ]
        assigned = assigned.reset_index()

    if assigned["Municipality Code"].isna().any():
        raise ValueError("Some 125 m cells could not be assigned to an administrative unit")
    if assigned["Mesh Code"].duplicated().any():
        raise ValueError("Municipality assignment must be one-to-one by Mesh Code")
    diagnostics = {
        "boundary_cells": float(len(boundary_codes)),
        "fallback_cells": float(len(unmatched_codes)),
        "maximum_fallback_distance_m": maximum_fallback_distance,
    }
    return assigned, diagnostics


def summarize_group(group: pd.DataFrame) -> pd.Series:
    """Calculate the pre-declared municipal reporting metrics."""
    population = pd.to_numeric(group["Total Population"], errors="coerce").fillna(0)
    population_total = float(population.sum())
    normal = pd.to_numeric(group["Normal Response Time (min)"], errors="coerce")
    disrupted = pd.to_numeric(group["Disrupted Response Time (min)"], errors="coerce")
    added = np.maximum(disrupted - normal, 0)
    backup = pd.to_numeric(group["Backup Fire Base Count"], errors="coerce")
    route_dependence = pd.to_numeric(
        group["Single Route Dependence"], errors="coerce"
    )
    route_assessed = route_dependence.notna()

    return pd.Series(
        {
            "Total Population": int(round(population_total)),
            "Population Age 65+ Share (%)": 100
            * weighted_mean(group["Population Age 65+ Share"], population),
            "Median Normal Response Time (min)": float(normal.median()),
            "Median Disrupted Response Time (min)": float(disrupted.median()),
            "Median Added Response Time (min)": float(pd.Series(added).median()),
            "Population within 10 min after Disruption (%)": 100
            * weighted_mean((disrupted <= TIMELY_RESPONSE_THRESHOLD_MIN).astype(float), population),
            "Population with No Backup Base within 10 min (%)": 100
            * weighted_mean((backup <= 0).astype(float), population),
            "Mean Single-Route Dependence among Assessed Cells (%)": 100
            * float(route_dependence.loc[route_assessed].mean())
            if route_assessed.any()
            else float("nan"),
            "High-Consequence 125 m Cells": int(
                (group["Population Conditional Fire Consequence Rank"] >= HIGH_CONSEQUENCE_THRESHOLD).sum()
            ),
            "Aggregate Population Conditional Fire Consequence": float(
                group["Population Conditional Fire Consequence"].sum()
            ),
        }
    )


def build_table() -> tuple[pd.DataFrame, dict[str, float]]:
    """Load, assign, and aggregate the accepted cell-level analysis results."""
    cells = gpd.read_parquet(CONSEQUENCE_PATH).to_crs(PROJECTED_CRS)
    administrative = gpd.read_parquet(ADMIN_PATH).to_crs(PROJECTED_CRS)
    cells["Mesh Code"] = cells["Mesh Code"].astype("string")
    administrative["Municipality Code"] = administrative["Municipality Code"].astype("string")
    assignment, diagnostics = assign_municipalities(cells, administrative)
    analysis = cells.drop(columns="Geometry").merge(
        assignment,
        on="Mesh Code",
        how="left",
        validate="one_to_one",
    )
    summary = (
        analysis.groupby(
            ["Municipality Code", "Municipality Label"],
            observed=True,
            sort=False,
        )
        .apply(summarize_group, include_groups=False)
        .reset_index()
    )
    summary = summary.sort_values(
        "Aggregate Population Conditional Fire Consequence",
        ascending=False,
        kind="stable",
    ).reset_index(drop=True)
    summary["Municipality"] = summary["Municipality Code"].map(ENGLISH_MUNICIPALITY)
    if summary["Municipality"].isna().any():
        missing_codes = summary.loc[
            summary["Municipality"].isna(), "Municipality Code"
        ].tolist()
        raise ValueError(f"Missing English municipality names for codes: {missing_codes}")
    summary = summary.drop(columns="Municipality Label").rename(
        columns={"Municipality": "Municipality Label"}
    )
    ordered_columns = [
        "Municipality Code",
        "Municipality Label",
        *[
            column
            for column in summary.columns
            if column not in {"Municipality Code", "Municipality Label"}
        ],
    ]
    summary = summary[ordered_columns]
    if len(summary) != len(administrative):
        raise ValueError(
            f"Expected {len(administrative)} administrative rows, received {len(summary)}"
        )
    if int(summary["Total Population"].sum()) != int(cells["Total Population"].sum()):
        raise ValueError("Municipality aggregation did not preserve total population")
    return summary, diagnostics


def build_notes(diagnostics: dict[str, float]) -> pd.DataFrame:
    """Document calculation and interpretation boundaries outside the main table."""
    notes = [
        (
            "Scope",
            "All results are planning screens conditional on a post-earthquake ignition; they are not fire probabilities, observed response times, or real-time capacity estimates.",
        ),
        (
            "Administrative assignment",
            "Each 125 m cell is assigned to the municipality or ward with the largest overlapping area; non-intersecting coastal slivers use the nearest administrative polygon.",
        ),
        (
            "Age 65+ share",
            "Population-weighted mean of disclosure-group age shares across assigned populated cells.",
        ),
        (
            "Response times",
            "Cell medians under the nominal normal-road network and the declared event-specific disruption rule; unmet service is capped at 30 minutes.",
        ),
        (
            "No backup base",
            "Population share in cells with no additional candidate dispatch base reachable within 10 minutes under event-specific disruption.",
        ),
        (
            "Single-route dependence",
            "Mean percentage of qualifying-base shortest routes that share the most-used internal road edge. A value of 100% indicates complete concentration. NA means that no cell in the reporting unit had a qualifying route for this assessment.",
        ),
        (
            "High-consequence cells",
            "Count of assigned cells at or above the prefecture-wide 90th percentile of Population Conditional Fire Consequence.",
        ),
        (
            "Aggregate consequence",
            "Sum of the relative cell-level Population Conditional Fire Consequence score; useful for within-study comparison, not a probability, burned area, or monetary loss.",
        ),
        (
            "Nearest-boundary fallback",
            f"{int(diagnostics['boundary_cells'])} cells required boundary review; {int(diagnostics['fallback_cells'])} non-intersecting cells used nearest-polygon assignment, with a maximum representative-point distance of {diagnostics['maximum_fallback_distance_m']:.1f} m.",
        ),
    ]
    return pd.DataFrame(notes, columns=["Item", "Definition or Boundary"])


def format_workbook(path: Path) -> None:
    """Apply consistent research-table formatting and numeric formats."""
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="263746")
    band_fill = PatternFill("solid", fgColor="F1F4F6")

    sheet = workbook["Municipality Summary"]
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    widths = [16, 20, 16, 17, 18, 20, 18, 20, 22, 22, 18, 23]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row_number in range(2, sheet.max_row + 1):
        if row_number % 2 == 0:
            for cell in sheet[row_number]:
                cell.fill = band_fill
        for column, cell in enumerate(sheet[row_number], start=1):
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="center",
                horizontal="left" if column == 2 else "center",
            )
        sheet.cell(row_number, 3).number_format = "#,##0"
        for column in range(4, 11):
            sheet.cell(row_number, column).number_format = "0.0"
        if sheet.cell(row_number, 10).value is None:
            sheet.cell(row_number, 10).value = "NA"
            sheet.cell(row_number, 10).number_format = "General"
        sheet.cell(row_number, 11).number_format = "0"
        sheet.cell(row_number, 12).number_format = "0.000"
        sheet.row_dimensions[row_number].height = 30
    sheet.row_dimensions[1].height = 58
    sheet.freeze_panes = "C2"
    sheet.auto_filter.ref = sheet.dimensions

    notes = workbook["Definitions"]
    for cell in notes[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    notes.column_dimensions["A"].width = 30
    notes.column_dimensions["B"].width = 110
    for row_number in range(2, notes.max_row + 1):
        notes.cell(row_number, 1).font = Font(bold=True, color="263746")
        notes.cell(row_number, 1).alignment = Alignment(vertical="top")
        notes.cell(row_number, 2).alignment = Alignment(wrap_text=True, vertical="top")
        notes.row_dimensions[row_number].height = 38
    notes.freeze_panes = "A2"
    notes.auto_filter.ref = notes.dimensions
    workbook.save(path)


def main() -> None:
    table, diagnostics = build_table()
    if table.shape != (49, 12):
        raise ValueError(f"Expected a 49 x 12 summary, received {table.shape}")
    notes = build_notes(diagnostics)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        table.to_excel(writer, sheet_name="Municipality Summary", index=False)
        notes.to_excel(writer, sheet_name="Definitions", index=False)
    format_workbook(OUTPUT)
    print(f"Saved: {OUTPUT.relative_to(ROOT)}")
    print(
        "Diagnostics: "
        f"rows={len(table)}; columns={len(table.columns)}; "
        f"population={int(table['Total Population'].sum()):,}; "
        f"boundary_cells={int(diagnostics['boundary_cells'])}; "
        f"fallback_cells={int(diagnostics['fallback_cells'])}; "
        f"maximum_fallback_distance={diagnostics['maximum_fallback_distance_m']:.1f} m"
    )


if __name__ == "__main__":
    main()
