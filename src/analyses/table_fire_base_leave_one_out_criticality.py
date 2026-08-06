#!/usr/bin/env python3
"""Fire Base Leave-One-Out Criticality.

Plan: report the 20 largest event-road population-objective losses after
removing one eligible fire base from the full represented response system.
Framework: AnaSOP Sections 5.3, 6.6, and workflow step 6.
"""

from __future__ import annotations

from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

from fire_base_criticality_common import load_fire_base_criticality
from fire_base_labels import fire_base_label_table


ROOT = Path(__file__).resolve().parents[2]
OUTPUT = ROOT / "data/results/tables/Table_fire_base_leave_one_out_criticality.xlsx"
OBJECTIVE = "Population"
TOP_BASE_COUNT = 20
HAN_PATTERN = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff]")


def build_table() -> tuple[pd.DataFrame, dict[str, int]]:
    values = load_fire_base_criticality()
    population = values.loc[
        values["Exposure Objective"].eq(OBJECTIVE)
        & values["Candidate Dispatch Base"].fillna(False)
    ].copy()
    if set(population["Road Scenario"].unique()) != {"normal", "central"}:
        raise ValueError("Expected normal and event-specific road scenarios")

    labels = fire_base_label_table(population)
    population = population.merge(
        labels,
        on="Fire Base Name",
        how="left",
        validate="many_to_one",
    )
    event = (
        population.loc[population["Road Scenario"].eq("central")]
        .sort_values(
            ["Leave-One-Out Fire Base Value Share", "Fire Base Name"],
            ascending=[False, True],
            kind="stable",
        )
        .head(TOP_BASE_COUNT)
        .set_index("Fire Base Name")
    )
    normal = population.loc[population["Road Scenario"].eq("normal")].set_index(
        "Fire Base Name"
    )
    normal_share = normal.reindex(event.index)["Leave-One-Out Fire Base Value Share"]
    event_share = event["Leave-One-Out Fire Base Value Share"]
    change_pp = 100 * (event_share - normal_share)
    effect = change_pp.map(lambda value: "↑" if value > 0 else ("↓" if value < 0 else "="))

    table = pd.DataFrame(
        {
            "Event-Road Order": range(1, TOP_BASE_COUNT + 1),
            "Fire Base": event["Fire Base"],
            "Base Type": event["Fire Base Type"],
            "Municipality": event["Municipality"],
            "Normal LOO Share (%)": 100 * normal_share,
            "Event LOO Share (%)": 100 * event_share,
            "Event − Normal (pp)": change_pp,
            "Two-Scenario IQR (pp)": 100
            * event["Road-Scenario Leave-One-Out IQR"],
            "Road Effect": effect,
        }
    ).reset_index(drop=True)
    diagnostics = {
        "eligible_bases": int(population["Fire Base Name"].nunique()),
        "verified_labels": int(event["Label Status"].eq("Verified project translation").sum()),
    }
    return table, diagnostics


def build_definitions(diagnostics: dict[str, int]) -> pd.DataFrame:
    rows = [
        (
            "Selection",
            f"The 20 largest event-road leave-one-out loss shares among {diagnostics['eligible_bases']} eligible candidate dispatch bases for the population objective.",
        ),
        (
            "LOO share (%)",
            "A base's increase in weighted accessibility loss after removing it from the full represented response system, divided by the summed removal losses across all eligible bases. This is not a Shapley or SHAP value.",
        ),
        (
            "Event − Normal (pp)",
            "Event-road LOO share minus normal-road LOO share, expressed in percentage points.",
        ),
        (
            "Two-scenario IQR (pp)",
            "The interquartile range across the normal-road and event-road LOO shares. With two declared scenarios, it is a compact descriptive contrast rather than a sampling uncertainty interval.",
        ),
        (
            "Road Effect",
            "↑ = larger dependence under event roads; ↓ = smaller dependence under event roads; = = no numerical change.",
        ),
        (
            "Interpretation",
            "Larger values indicate greater modelled system dependence on that base because alternatives substitute less effectively when it is removed.",
        ),
        (
            "Boundary",
            "The results do not measure actual vehicles, staffing, dispatch decisions, station quality, observed fire losses, or causal effects.",
        ),
        (
            "English labels",
            f"All {diagnostics['verified_labels']} displayed fire-base labels use verified project translations; source-language names remain available in derived data.",
        ),
    ]
    return pd.DataFrame(rows, columns=["Item", "Definition or Boundary"])


def format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="263746")
    band_fill = PatternFill("solid", fgColor="F1F4F6")

    sheet = workbook["Top 20 Fire Bases"]
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    widths = [14, 40, 20, 24, 18, 18, 19, 20, 14]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row_number in range(2, sheet.max_row + 1):
        if row_number % 2 == 0:
            for cell in sheet[row_number]:
                cell.fill = band_fill
        for column, cell in enumerate(sheet[row_number], start=1):
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="center",
                horizontal="left" if column in {2, 3, 4} else "center",
            )
        for column in range(5, 9):
            sheet.cell(row_number, column).number_format = "0.000"
        sheet.cell(row_number, 9).font = Font(bold=True, size=12, color="263746")
        sheet.row_dimensions[row_number].height = 34
    sheet.row_dimensions[1].height = 52
    sheet.freeze_panes = "E2"
    sheet.auto_filter.ref = sheet.dimensions

    definitions = workbook["Definitions"]
    for cell in definitions[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    definitions.column_dimensions["A"].width = 29
    definitions.column_dimensions["B"].width = 112
    for row_number in range(2, definitions.max_row + 1):
        definitions.cell(row_number, 1).font = Font(bold=True, color="263746")
        definitions.cell(row_number, 1).alignment = Alignment(vertical="top")
        definitions.cell(row_number, 2).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
        definitions.row_dimensions[row_number].height = 44
    definitions.freeze_panes = "A2"
    workbook.save(path)


def validate_workbook(path: Path) -> None:
    workbook = load_workbook(path, data_only=True)
    hits = [
        (sheet.title, cell.coordinate, cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and HAN_PATTERN.search(cell.value)
    ]
    if hits:
        raise ValueError(f"Han characters remain in workbook: {hits[:5]}")


def main() -> None:
    table, diagnostics = build_table()
    if table.shape != (TOP_BASE_COUNT, 9):
        raise ValueError(f"Expected a 20 x 9 table, received {table.shape}")
    if not table["Event LOO Share (%)"].is_monotonic_decreasing:
        raise ValueError("Event-road criticality order is not descending")
    definitions = build_definitions(diagnostics)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        table.to_excel(writer, sheet_name="Top 20 Fire Bases", index=False)
        definitions.to_excel(writer, sheet_name="Definitions", index=False)
    format_workbook(OUTPUT)
    validate_workbook(OUTPUT)
    print(f"Saved: {OUTPUT.relative_to(ROOT)}")
    print(
        "Diagnostics: "
        f"rows={len(table)}; columns={len(table.columns)}; "
        f"eligible_bases={diagnostics['eligible_bases']}; "
        f"verified_labels={diagnostics['verified_labels']}; "
        "han_character_cells=0"
    )


if __name__ == "__main__":
    main()
