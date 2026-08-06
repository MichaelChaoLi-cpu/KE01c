#!/usr/bin/env python3
"""Scenario Definitions and Interpretation Boundaries

Plan: catalogue the reference, event-specific, stochastic, one-base-removal, and
intervention scenarios together with their permitted interpretations.
Framework: AnaSOP Sections 5, 6.1-6.8, and workflow steps 3-8.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd


ROOT = Path(__file__).resolve().parents[2]
OUTPUT = ROOT / "data/results/tables/Table_scenario_definitions_and_interpretation_boundaries.xlsx"

CONDITIONAL_IGNITION = (
    "Condition on a post-earthquake ignition; no ignition location or probability is predicted"
)
NORMAL_WATER = "Normal-reliance boundary; hydrant status is not observed"
BOUNDED_WATER = (
    "Mapped Minami Ward supply interruption treated as a bounded non-reliance stress"
)
ELIGIBLE_BASES = "81 eligible candidate dispatch bases; network reachability varies by road state"


def build_table() -> pd.DataFrame:
    rows = [
        {
            "Scenario label": "Normal-road reference",
            "Imposed ignition condition": CONDITIONAL_IGNITION,
            "Event-specific road rule": "Retain every eligible routable road edge",
            "Road Failure Model": "None",
            "Expected Failed Road Length Share": "Not imposed",
            "Water constraint": NORMAL_WATER,
            "Available candidate bases": ELIGIBLE_BASES,
            "Analytical purpose": "Reference response time, redundancy, and base dependence",
            "Interpretation boundary": "Nominal network benchmark, not observed travel time or real-time capacity",
        },
        {
            "Scenario label": "Event-specific road disruption",
            "Imposed ignition condition": CONDITIONAL_IGNITION,
            "Event-specific road rule": "Remove eligible edges mapped in Warning or Special Warning exposure classes",
            "Road Failure Model": "None",
            "Expected Failed Road Length Share": "Not assigned",
            "Water constraint": NORMAL_WATER,
            "Available candidate bases": ELIGIBLE_BASES,
            "Analytical purpose": "Primary post-event accessibility and base-dependence contrast",
            "Interpretation boundary": "Transparent planning rule; not reconstructed 2026 road passability",
        },
        {
            "Scenario label": "Bounded water constraint",
            "Imposed ignition condition": CONDITIONAL_IGNITION,
            "Event-specific road rule": "Normal-road reference",
            "Road Failure Model": "None",
            "Expected Failed Road Length Share": "Not imposed",
            "Water constraint": BOUNDED_WATER,
            "Available candidate bases": ELIGIBLE_BASES,
            "Analytical purpose": "Bound the consequence of non-reliance on mapped water supply",
            "Interpretation boundary": "Does not identify individual hydrant functionality or water pressure",
        },
        {
            "Scenario label": "Combined event and water stress",
            "Imposed ignition condition": CONDITIONAL_IGNITION,
            "Event-specific road rule": "Apply the event-specific road-disruption rule",
            "Road Failure Model": "None",
            "Expected Failed Road Length Share": "Not assigned",
            "Water constraint": BOUNDED_WATER,
            "Available candidate bases": ELIGIBLE_BASES,
            "Analytical purpose": "Main conditional-consequence and intervention baseline",
            "Interpretation boundary": "Scenario screen, not a forecast of burned area or realized firefighting performance",
        },
        {
            "Scenario label": "Independent road-section sensitivity — 1%",
            "Imposed ignition condition": CONDITIONAL_IGNITION,
            "Event-specific road rule": "Event-specific disruption plus full junction-to-junction section closure",
            "Road Failure Model": "Length-dependent independent",
            "Expected Failed Road Length Share": "1%",
            "Water constraint": NORMAL_WATER,
            "Available candidate bases": ELIGIBLE_BASES,
            "Analytical purpose": "Low-severity stochastic reliability sensitivity",
            "Interpretation boundary": "Declared Monte Carlo severity, not an engineering failure probability",
        },
        {
            "Scenario label": "Independent formal reliability — 3%",
            "Imposed ignition condition": CONDITIONAL_IGNITION,
            "Event-specific road rule": "Event-specific disruption plus full junction-to-junction section closure",
            "Road Failure Model": "Length-dependent independent",
            "Expected Failed Road Length Share": "3%",
            "Water constraint": NORMAL_WATER,
            "Available candidate bases": ELIGIBLE_BASES,
            "Analytical purpose": "Primary 1,000-replicate timely-response reliability estimate",
            "Interpretation boundary": "Formal scenario estimate; still not observed or engineering-calibrated failure risk",
        },
        {
            "Scenario label": "Independent road-section sensitivity — 5%",
            "Imposed ignition condition": CONDITIONAL_IGNITION,
            "Event-specific road rule": "Event-specific disruption plus full junction-to-junction section closure",
            "Road Failure Model": "Length-dependent independent",
            "Expected Failed Road Length Share": "5%",
            "Water constraint": NORMAL_WATER,
            "Available candidate bases": ELIGIBLE_BASES,
            "Analytical purpose": "Moderate-severity stochastic reliability sensitivity",
            "Interpretation boundary": "100-replicate mechanism-comparison pilot, not the formal 1,000-replicate estimate",
        },
        {
            "Scenario label": "Independent stress test — 10%",
            "Imposed ignition condition": CONDITIONAL_IGNITION,
            "Event-specific road rule": "Event-specific disruption plus full junction-to-junction section closure",
            "Road Failure Model": "Length-dependent independent",
            "Expected Failed Road Length Share": "10%",
            "Water constraint": NORMAL_WATER,
            "Available candidate bases": ELIGIBLE_BASES,
            "Analytical purpose": "High-severity stress boundary",
            "Interpretation boundary": "Stress test only; not an expected 2026 road-loss level",
        },
        {
            "Scenario label": "Spatially clustered pilot — 3%",
            "Imposed ignition condition": CONDITIONAL_IGNITION,
            "Event-specific road rule": "Event-specific disruption plus correlated full-section closures within 5 km cells",
            "Road Failure Model": "Spatially clustered",
            "Expected Failed Road Length Share": "3%",
            "Water constraint": NORMAL_WATER,
            "Available candidate bases": ELIGIBLE_BASES,
            "Analytical purpose": "Test sensitivity to geographically concentrated loss",
            "Interpretation boundary": "100-replicate pilot with imposed spatial correlation, not an empirical damage field",
        },
        {
            "Scenario label": "Hazard-weighted pilot — 3%",
            "Imposed ignition condition": CONDITIONAL_IGNITION,
            "Event-specific road rule": "Event-specific disruption plus higher closure weight on event-exposed sections",
            "Road Failure Model": "Hazard-weighted",
            "Expected Failed Road Length Share": "3%",
            "Water constraint": NORMAL_WATER,
            "Available candidate bases": ELIGIBLE_BASES,
            "Analytical purpose": "Test sensitivity to preferential failure in mapped event-exposure areas",
            "Interpretation boundary": "100-replicate pilot; weighting is imposed rather than engineering-calibrated",
        },
        {
            "Scenario label": "One-fire-base removal",
            "Imposed ignition condition": CONDITIONAL_IGNITION,
            "Event-specific road rule": "Evaluate normal, event-specific, and declared stochastic road states",
            "Road Failure Model": "None or declared stochastic mechanism",
            "Expected Failed Road Length Share": "Not imposed or 1-10%",
            "Water constraint": NORMAL_WATER,
            "Available candidate bases": "80 of 81 eligible bases per run; each base is removed once in turn",
            "Analytical purpose": "Measure full-system accessibility loss when one base is removed",
            "Interpretation boundary": "System dependence, not intrinsic quality, observed capacity, staffing, dispatch, or causal effect",
        },
        {
            "Scenario label": "Resource intervention counterfactual",
            "Imposed ignition condition": CONDITIONAL_IGNITION,
            "Event-specific road rule": "Event-specific baseline with selected road sections restored",
            "Road Failure Model": "None",
            "Expected Failed Road Length Share": "Not assigned",
            "Water constraint": "Bounded water stress with selected 1 km support areas",
            "Available candidate bases": "Eligible bases plus selected mapped candidate staging sites",
            "Analytical purpose": "Compare within-class action bundles and simple baselines",
            "Interpretation boundary": "Relative screening benefit; staging sites require field verification, and cross-class bundles are not cost-optimal without comparable costs",
        },
    ]
    table = pd.DataFrame(rows)
    scenario = table["Scenario label"]
    road_code = {
        "Normal-road reference": "R0",
        "Event-specific road disruption": "RE",
        "Bounded water constraint": "R0",
        "Combined event and water stress": "RE",
        "Independent road-section sensitivity — 1%": "RE",
        "Independent formal reliability — 3%": "RE",
        "Independent road-section sensitivity — 5%": "RE",
        "Independent stress test — 10%": "RE",
        "Spatially clustered pilot — 3%": "RE",
        "Hazard-weighted pilot — 3%": "RE",
        "One-fire-base removal": "R0 / RE / RF",
        "Resource intervention counterfactual": "RE + RS",
    }
    failure_code = {
        "None": "—",
        "Length-dependent independent": "LD",
        "Spatially clustered": "SC",
        "Hazard-weighted": "HW",
        "None or declared stochastic mechanism": "— / LD / SC / HW",
    }
    water_code = {
        "Normal-road reference": "W0",
        "Event-specific road disruption": "W0",
        "Bounded water constraint": "WB",
        "Combined event and water stress": "WB",
        "Independent road-section sensitivity — 1%": "W0",
        "Independent formal reliability — 3%": "W0",
        "Independent road-section sensitivity — 5%": "W0",
        "Independent stress test — 10%": "W0",
        "Spatially clustered pilot — 3%": "W0",
        "Hazard-weighted pilot — 3%": "W0",
        "One-fire-base removal": "W0",
        "Resource intervention counterfactual": "WB + WS",
    }
    base_code = {
        "Normal-road reference": "B81",
        "Event-specific road disruption": "B81",
        "Bounded water constraint": "B81",
        "Combined event and water stress": "B81",
        "Independent road-section sensitivity — 1%": "B81",
        "Independent formal reliability — 3%": "B81",
        "Independent road-section sensitivity — 5%": "B81",
        "Independent stress test — 10%": "B81",
        "Spatially clustered pilot — 3%": "B81",
        "Hazard-weighted pilot — 3%": "B81",
        "One-fire-base removal": "B81−1",
        "Resource intervention counterfactual": "B81 + T",
    }
    use_code = {
        "Normal-road reference": "REF",
        "Event-specific road disruption": "ACC",
        "Bounded water constraint": "WAT",
        "Combined event and water stress": "CON",
        "Independent road-section sensitivity — 1%": "REL",
        "Independent formal reliability — 3%": "REL-F",
        "Independent road-section sensitivity — 5%": "REL",
        "Independent stress test — 10%": "STR",
        "Spatially clustered pilot — 3%": "ROB",
        "Hazard-weighted pilot — 3%": "ROB",
        "One-fire-base removal": "LOO",
        "Resource intervention counterfactual": "INT",
    }
    boundary_code = {
        "Normal-road reference": "NOM",
        "Event-specific road disruption": "EVT",
        "Bounded water constraint": "BND",
        "Combined event and water stress": "SCN",
        "Independent road-section sensitivity — 1%": "PIL",
        "Independent formal reliability — 3%": "FORM",
        "Independent road-section sensitivity — 5%": "PIL",
        "Independent stress test — 10%": "STR",
        "Spatially clustered pilot — 3%": "PIL",
        "Hazard-weighted pilot — 3%": "PIL",
        "One-fire-base removal": "DEP",
        "Resource intervention counterfactual": "CF",
    }
    nominal_road_input = table["Expected Failed Road Length Share"].replace(
        {
            "Not imposed": "—",
            "Not assigned": "—",
            "Not imposed or 1-10%": "— / 1–10% nominal",
        }
    )
    nominal_road_input = nominal_road_input.map(
        lambda value: f"{value} nominal" if str(value).endswith("%") else value
    )
    compact = pd.DataFrame(
        {
            "Scenario": scenario,
            "Road": scenario.map(road_code),
            "Failure": table["Road Failure Model"].map(failure_code),
            "Nominal road input": nominal_road_input,
            "Water": scenario.map(water_code),
            "Bases": scenario.map(base_code),
            "Use": scenario.map(use_code),
            "Boundary": scenario.map(boundary_code),
        }
    )
    symbol_key = pd.DataFrame(
        [
            {
                "Scenario": "How to read the symbols",
                "Road": "R0 normal roads; RE event disruption; RF declared stochastic road states; RS selected sections restored",
                "Failure": "LD independent; SC spatial cluster; HW hazard weighted; — none",
                "Nominal road input": (
                    "1%, 3%, 5%, or 10% = nominal expected failed-length input over the "
                    "eligible pre-mask network; mechanisms are stress scenarios, not equal-net-damage comparisons"
                ),
                "Water": "W0 normal reliance; WB bounded disruption; WS temporary support",
                "Bases": "B81 all 81 bases; B81−1 one base removed in turn; T temporary origins",
                "Use": (
                    "REF reference; ACC accessibility; WAT water; CON combined; REL sensitivity; "
                    "REL-F formal estimate; STR stress test; ROB robustness; LOO one-base removal; INT intervention"
                ),
                "Boundary": (
                    "NOM nominal; EVT event rule; BND bounded assumption; SCN scenario; PIL pilot; "
                    "FORM formal estimate; DEP system dependence; CF counterfactual"
                ),
            }
        ]
    )
    compact = pd.concat([compact, symbol_key], ignore_index=True)
    if compact.isna().any().any():
        raise ValueError("A compact scenario code is missing")
    return compact


def format_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    sheet = workbook["Scenarios"]
    header_fill = PatternFill("solid", fgColor="263746")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    widths = [38, 14, 18, 23, 16, 16, 18, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for index, cell in enumerate(row, start=1):
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="center",
                horizontal="left" if index == 1 else "center",
            )
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 34
    for index in range(2, sheet.max_row):
        sheet.row_dimensions[index].height = 30
    key_fill = PatternFill("solid", fgColor="E8EEF2")
    for cell in sheet[sheet.max_row]:
        cell.fill = key_fill
        cell.font = Font(color="263746", bold=cell.column == 1, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    sheet.row_dimensions[sheet.max_row].height = 118
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_title_rows = "1:1"
    sheet.print_area = sheet.dimensions
    sheet.page_margins.left = 0.2
    sheet.page_margins.right = 0.2
    sheet.page_margins.top = 0.3
    sheet.page_margins.bottom = 0.3
    workbook.save(path)


def main() -> None:
    table = build_table()
    if table.shape != (13, 8):
        raise ValueError(f"Expected a 13 x 8 scenario table, received {table.shape}")
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        table.to_excel(writer, sheet_name="Scenarios", index=False)
    format_workbook(OUTPUT)
    print(f"Saved: {OUTPUT.relative_to(ROOT)}")
    print(f"Diagnostics: scenario_rows={len(table) - 1}; columns={len(table.columns)}; key_rows=1")


if __name__ == "__main__":
    main()
