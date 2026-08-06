#!/usr/bin/env python3
"""Generate a compact paired intervention-performance table by budget."""

from __future__ import annotations

from pathlib import Path
import re

import numpy as np
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd


ROOT = Path(__file__).resolve().parents[2]
PERFORMANCE_PATH = ROOT / "data/results/derived/intervention_performance.parquet"
OUTPUT = ROOT / "data/results/tables/Table_intervention_performance_by_budget.xlsx"

DISPLAY_BUDGETS = (1, 3, 5)
# Retain the stored strategy label for compatibility with the derived data,
# while presenting the method to readers as prioritized selection.
PRIORITIZED_SOURCE = "Greedy consequence reduction"
BASELINE = "Simple baseline"
HAN_PATTERN = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff]")


def build_table() -> tuple[pd.DataFrame, dict[str, float]]:
    """Pair prioritized and baseline estimates at representative budgets."""
    performance = pd.read_parquet(PERFORMANCE_PATH)
    selected = performance.loc[performance["Budget"].isin(DISPLAY_BUDGETS)].copy()
    index = ["Action Type", "Budget Definition", "Budget"]
    value_columns = ["Budget Used", "Intervention Benefit", "Consequence Reduction Share"]
    paired = selected.pivot(index=index, columns="Strategy", values=value_columns)
    expected_strategies = {PRIORITIZED_SOURCE, BASELINE}
    if set(selected["Strategy"].unique()) != expected_strategies:
        raise ValueError("Expected prioritized-source and simple-baseline strategies")

    paired.columns = [f"{metric}::{strategy}" for metric, strategy in paired.columns]
    paired = paired.reset_index()
    prioritized_benefit = paired[f"Intervention Benefit::{PRIORITIZED_SOURCE}"]
    baseline_benefit = paired[f"Intervention Benefit::{BASELINE}"]
    advantage = np.where(
        baseline_benefit > 0,
        100 * (prioritized_benefit - baseline_benefit) / baseline_benefit,
        np.nan,
    )
    table = pd.DataFrame(
        {
            "Action Type": paired["Action Type"],
            "Budget Definition": paired["Budget Definition"],
            "Budget": paired["Budget"],
            "Prioritized Budget Used": paired[f"Budget Used::{PRIORITIZED_SOURCE}"],
            "Baseline Budget Used": paired[f"Budget Used::{BASELINE}"],
            "Prioritized Intervention Benefit": prioritized_benefit,
            "Baseline Intervention Benefit": baseline_benefit,
            "Prioritized Consequence Reduction (%)": 100
            * paired[f"Consequence Reduction Share::{PRIORITIZED_SOURCE}"],
            "Baseline Consequence Reduction (%)": 100
            * paired[f"Consequence Reduction Share::{BASELINE}"],
            "Prioritized Advantage over Baseline (%)": advantage,
        }
    )
    action_order = {
        "Candidate staging site": 0,
        "Bounded water support": 1,
        "Priority road restoration": 2,
    }
    budget_order = {
        "Action count": 0,
        "Road section count": 1,
        "Normalized event-exposed length": 2,
    }
    table = (
        table.assign(
            _action=table["Action Type"].map(action_order),
            _definition=table["Budget Definition"].map(budget_order),
        )
        .sort_values(["_action", "_definition", "Budget"], kind="stable")
        .drop(columns=["_action", "_definition"])
        .reset_index(drop=True)
    )
    diagnostics = {
        "full_rows": float(len(performance)),
        "baseline_loss": float(performance["Baseline Combined-Stress Loss"].iloc[0]),
        "zero_baseline_rows": float((baseline_benefit <= 0).sum()),
        "prioritized_wins": float((prioritized_benefit > baseline_benefit).sum()),
        "ties": float(np.isclose(prioritized_benefit, baseline_benefit).sum()),
    }
    return table, diagnostics


def build_definitions(diagnostics: dict[str, float]) -> pd.DataFrame:
    """Document row selection, units, comparators, and interpretation limits."""
    rows = [
        (
            "Displayed budgets",
            f"Budgets 1, 3, and 5 are shown for each of four action and budget-definition paths, reducing the reader-facing table from {int(diagnostics['full_rows'])} strategy rows to 12 paired rows. Complete budgets 1 through 5 remain in derived data.",
        ),
        (
            "Event scenario",
            "All rows use the same event-specific road and bounded-water combined-stress baseline, so the invariant scenario label is not repeated in the main table.",
        ),
        (
            "Prioritized strategy",
            "Within-class forward selection using the complete modelled conditional-consequence objective. Action-count and road-section-count paths maximize exact marginal reduction; the normalized event-exposed-length path recomputes exact marginal reduction per cost after every selected road section.",
        ),
        (
            "Simple baseline",
            "Population-oriented selection among the same traceable candidate staging sites and water-support areas; road-class, emergency-route, and bridge-screening priority for road restoration.",
        ),
        (
            "Budget used",
            "Number of actions for action-count and road-section-count rows; accumulated Road Restoration Cost Proxy for normalized event-exposed-length rows. Units are not comparable across action classes.",
        ),
        (
            "Intervention benefit",
            f"Reduction in modelled system loss relative to the fixed combined-stress baseline of {diagnostics['baseline_loss']:.3f} score units.",
        ),
        (
            "Consequence reduction",
            "Intervention Benefit divided by baseline combined-stress loss, expressed as a percentage.",
        ),
        (
            "Prioritized advantage",
            "Percentage improvement of prioritized benefit over the paired simple baseline. NA means the paired baseline benefit is zero, so a relative percentage is undefined.",
        ),
        (
            "Interpretation boundary",
            "These are within-class scenario comparisons, not observed outcomes or a cost-optimal mixed portfolio. Candidate staging sites are mapped public or emergency locations that still require field verification; cross-class budget units are not commensurable.",
        ),
    ]
    return pd.DataFrame(rows, columns=["Item", "Definition or Boundary"])


def format_workbook(path: Path) -> None:
    """Apply compact formatting to the 12 x 10 paired table."""
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="263746")
    band_fill = PatternFill("solid", fgColor="F1F4F6")
    sheet = workbook["Budget Performance"]
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    widths = [27, 30, 12, 18, 18, 21, 21, 21, 21, 22]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row_number in range(2, sheet.max_row + 1):
        if row_number % 2 == 0:
            for cell in sheet[row_number]:
                cell.fill = band_fill
        for column, cell in enumerate(sheet[row_number], start=1):
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="center",
                horizontal="left" if column in {1, 2} else "center",
            )
        for column in range(4, 10):
            sheet.cell(row_number, column).number_format = "0.000"
        if sheet.cell(row_number, 10).value is None:
            sheet.cell(row_number, 10).value = "NA"
            sheet.cell(row_number, 10).number_format = "General"
        else:
            sheet.cell(row_number, 10).number_format = "0.0"
        sheet.row_dimensions[row_number].height = 34
    sheet.row_dimensions[1].height = 64
    sheet.freeze_panes = "D2"
    sheet.auto_filter.ref = sheet.dimensions

    definitions = workbook["Definitions"]
    for cell in definitions[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    definitions.column_dimensions["A"].width = 31
    definitions.column_dimensions["B"].width = 112
    for row_number in range(2, definitions.max_row + 1):
        definitions.cell(row_number, 1).font = Font(bold=True, color="263746")
        definitions.cell(row_number, 1).alignment = Alignment(vertical="top")
        definitions.cell(row_number, 2).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
        definitions.row_dimensions[row_number].height = 44
    definitions.freeze_panes = "A2"
    workbook.save(path)


def validate_english_only(path: Path) -> None:
    """Fail if any reader-facing workbook cell contains a Han character."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    hits = [
        (sheet.title, cell.coordinate, cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and HAN_PATTERN.search(cell.value)
    ]
    if hits:
        raise ValueError(f"Han characters remain in workbook: {hits[:5]}")


def main() -> None:
    table, diagnostics = build_table()
    if table.shape != (12, 10):
        raise ValueError(f"Expected a 12 x 10 table, received {table.shape}")
    definitions = build_definitions(diagnostics)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        table.to_excel(writer, sheet_name="Budget Performance", index=False)
        definitions.to_excel(writer, sheet_name="Definitions", index=False)
    format_workbook(OUTPUT)
    validate_english_only(OUTPUT)
    print(f"Saved: {OUTPUT.relative_to(ROOT)}")
    print(
        "Diagnostics: "
        f"rows={len(table)}; columns={len(table.columns)}; "
        f"prioritized_wins={int(diagnostics['prioritized_wins'])}; "
        f"ties={int(diagnostics['ties'])}; "
        f"zero_baseline_rows={int(diagnostics['zero_baseline_rows'])}; "
        "han_character_cells=0"
    )


if __name__ == "__main__":
    main()
