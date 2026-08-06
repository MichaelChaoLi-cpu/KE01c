#!/usr/bin/env python3
"""Generate a compact English-only table of priority critical facilities."""

from __future__ import annotations

from pathlib import Path
import re

import geopandas as gpd
import numpy as np
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

from table_municipality_fire_consequence_and_accessibility_summary import (
    ENGLISH_MUNICIPALITY,
    assign_municipalities,
)


ROOT = Path(__file__).resolve().parents[2]
PROCESSED = ROOT / "data/processed"
CONTEXT_PATH = ROOT / "data/results/derived/intervention_context_125m.parquet"
ADMIN_PATH = PROCESSED / "administrative_areas_preprocessed.parquet"
OUTPUT = ROOT / "data/results/tables/Table_priority_critical_facilities.xlsx"

PROJECTED_CRS = 6670
PER_CLASS_COUNT = 4
FACILITY_CLASS_COUNT = 5
OUTPUT_ROWS = PER_CLASS_COUNT * FACILITY_CLASS_COUNT
NEAREST_CELL_LIMIT_M = 250.0
HAN_PATTERN = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff]")


def readable_value(value: object) -> str:
    """Return a compact English-safe representation of a source code or count."""
    if pd.isna(value):
        return "NA"
    if isinstance(value, (float, np.floating)) and float(value).is_integer():
        return str(int(value))
    return str(value).strip() or "NA"


def facility_layers() -> gpd.GeoDataFrame:
    """Harmonize five facility classes without carrying Japanese names forward."""
    medical = gpd.read_parquet(PROCESSED / "medical_facilities_preprocessed.parquet")
    medical_frame = medical[["Medical Facility ID", "Geometry"]].rename(
        columns={"Medical Facility ID": "Facility ID"}
    )
    medical_frame["Facility Class"] = "Medical facility"
    medical_frame["Capacity or Designation"] = [
        f"Beds {readable_value(beds)}; emergency code {readable_value(emergency)}; disaster-base code {readable_value(disaster)}"
        for beds, emergency, disaster in zip(
            medical["Bed Count"],
            medical["Emergency Hospital Designation"],
            medical["Disaster Base Hospital Class"],
            strict=True,
        )
    ]

    welfare = gpd.read_parquet(PROCESSED / "welfare_facilities_preprocessed.parquet")
    welfare_frame = welfare[["Welfare Facility ID", "Geometry"]].rename(
        columns={"Welfare Facility ID": "Facility ID"}
    )
    welfare_frame["Facility Class"] = "Welfare facility"
    welfare_frame["Capacity or Designation"] = [
        f"Welfare class code {readable_value(code)}"
        for code in welfare["Welfare Facility Minor Class"]
    ]

    schools = gpd.read_parquet(PROCESSED / "schools_preprocessed.parquet")
    school_frame = schools[["School Facility ID", "Geometry"]].rename(
        columns={"School Facility ID": "Facility ID"}
    )
    school_frame["Facility Class"] = "School"
    school_frame["Capacity or Designation"] = [
        f"School class code {readable_value(code)}; suspension code {readable_value(status)}"
        for code, status in zip(
            schools["School Class"],
            schools["Suspension Status"],
            strict=True,
        )
    ]

    shelters = gpd.read_parquet(PROCESSED / "designated_shelters_preprocessed.parquet")
    shelter_frame = shelters[["Shelter ID", "Geometry"]].rename(
        columns={"Shelter ID": "Facility ID"}
    )
    shelter_frame["Facility Class"] = "Designated shelter"
    shelter_frame["Capacity or Designation"] = [
        f"Accepted persons {readable_value(capacity)}"
        for capacity in shelters["Accepted Persons"]
    ]

    evacuation = gpd.read_parquet(
        PROCESSED / "emergency_evacuation_sites_preprocessed.parquet"
    )
    evacuation_frame = evacuation[["Evacuation Site ID", "Geometry"]].rename(
        columns={"Evacuation Site ID": "Facility ID"}
    )
    evacuation_frame["Facility Class"] = "Emergency evacuation site"
    evacuation_frame["Capacity or Designation"] = [
        f"Earthquake code {readable_value(earthquake)}; large-fire code {readable_value(fire)}"
        for earthquake, fire in zip(
            evacuation["Earthquake Designation"],
            evacuation["Large-Scale Fire Designation"],
            strict=True,
        )
    ]

    combined = pd.concat(
        [
            medical_frame,
            welfare_frame,
            school_frame,
            shelter_frame,
            evacuation_frame,
        ],
        ignore_index=True,
    )
    facilities = gpd.GeoDataFrame(combined, geometry="Geometry", crs=medical.crs)
    facilities = facilities.loc[
        facilities["Facility ID"].notna() & facilities.geometry.notna()
    ].copy()
    facilities["Facility ID"] = facilities["Facility ID"].astype("string")
    facilities["Facility Key"] = (
        facilities["Facility Class"].astype("string")
        + "::"
        + facilities["Facility ID"].astype("string")
    )
    if facilities["Facility Key"].duplicated().any():
        raise ValueError("Facility class and ID must form a unique key")
    return facilities.to_crs(PROJECTED_CRS)


def attach_cells(
    facilities: gpd.GeoDataFrame,
    cells: gpd.GeoDataFrame,
) -> tuple[gpd.GeoDataFrame, dict[str, float]]:
    """Assign facilities to modeled cells, with a bounded nearest-cell fallback."""
    cell_columns = [
        "Mesh Code",
        "Combined Stress Conditional Consequence",
        "Combined Stress Consequence Rank",
        "Accessibility Penalty",
        "Accessibility Weakness Rank",
        "Disrupted Response Time (min)",
        "Backup Fire Base Count",
        "Municipality Code",
        "Municipality",
        "Geometry",
    ]
    modeled = cells[cell_columns].copy()
    matched = gpd.sjoin(
        facilities,
        modeled,
        how="left",
        predicate="within",
    ).drop(columns="index_right")
    matched = matched.sort_values(["Facility Key", "Mesh Code"], kind="stable").drop_duplicates(
        "Facility Key"
    )
    unmatched_keys = matched.loc[matched["Mesh Code"].isna(), "Facility Key"]
    fallback_count = 0
    maximum_distance = 0.0
    if len(unmatched_keys):
        unmatched = facilities.loc[facilities["Facility Key"].isin(unmatched_keys)].copy()
        nearest = gpd.sjoin_nearest(
            unmatched,
            modeled,
            how="left",
            max_distance=NEAREST_CELL_LIMIT_M,
            distance_col="Cell Distance (m)",
        ).drop(columns="index_right")
        nearest = nearest.sort_values(
            ["Facility Key", "Cell Distance (m)", "Mesh Code"],
            kind="stable",
        ).drop_duplicates("Facility Key")
        usable = nearest.loc[nearest["Mesh Code"].notna()].copy()
        fallback_count = len(usable)
        if fallback_count:
            maximum_distance = float(usable["Cell Distance (m)"].max())
            matched = matched.set_index("Facility Key")
            usable = usable.set_index("Facility Key")
            update_columns = [column for column in cell_columns if column != "Geometry"]
            matched.loc[usable.index, update_columns] = usable[update_columns]
            matched = matched.reset_index()
    diagnostics = {
        "source_facilities": float(len(facilities)),
        "assigned_facilities": float(matched["Mesh Code"].notna().sum()),
        "fallback_facilities": float(fallback_count),
        "maximum_fallback_distance_m": maximum_distance,
    }
    return matched.loc[matched["Mesh Code"].notna()].copy(), diagnostics


def build_table() -> tuple[pd.DataFrame, pd.DataFrame, dict[str, float]]:
    """Rank facilities with an OR-style consequence/accessibility priority score."""
    context = gpd.read_parquet(CONTEXT_PATH).to_crs(PROJECTED_CRS)
    context["Mesh Code"] = context["Mesh Code"].astype("string")
    administrative = gpd.read_parquet(ADMIN_PATH).to_crs(PROJECTED_CRS)
    administrative["Municipality Code"] = administrative["Municipality Code"].astype(
        "string"
    )
    assignment, admin_diagnostics = assign_municipalities(context, administrative)
    assignment["Municipality"] = assignment["Municipality Code"].map(
        ENGLISH_MUNICIPALITY
    )
    if assignment["Municipality"].isna().any():
        raise ValueError("English municipality assignment is incomplete")
    context = context.merge(
        assignment[["Mesh Code", "Municipality Code", "Municipality"]],
        on="Mesh Code",
        how="left",
        validate="one_to_one",
    )
    context["Accessibility Weakness Rank"] = context["Accessibility Penalty"].rank(
        method="average",
        pct=True,
    )

    facilities = facility_layers()
    attached, facility_diagnostics = attach_cells(facilities, context)
    attached["Facility Priority Score"] = attached[
        ["Combined Stress Consequence Rank", "Accessibility Weakness Rank"]
    ].max(axis=1)
    attached = attached.sort_values(
        [
            "Facility Class",
            "Facility Priority Score",
            "Combined Stress Conditional Consequence",
            "Disrupted Response Time (min)",
            "Facility ID",
        ],
        ascending=[True, False, False, False, True],
        kind="stable",
    )
    selected = attached.groupby("Facility Class", sort=False).head(PER_CLASS_COUNT).copy()
    if selected["Facility Class"].nunique() != FACILITY_CLASS_COUNT:
        raise ValueError("All five facility classes must be represented")
    selected = selected.sort_values(
        [
            "Facility Priority Score",
            "Combined Stress Conditional Consequence",
            "Facility Class",
            "Facility ID",
        ],
        ascending=[False, False, True, True],
        kind="stable",
    ).reset_index(drop=True)
    selected["Priority Rank"] = np.arange(1, len(selected) + 1, dtype=int)

    table = selected[
        [
            "Priority Rank",
            "Facility Class",
            "Municipality",
            "Capacity or Designation",
            "Mesh Code",
            "Combined Stress Conditional Consequence",
            "Disrupted Response Time (min)",
            "Backup Fire Base Count",
            "Facility Priority Score",
        ]
    ].rename(
        columns={
            "Combined Stress Conditional Consequence": "Combined-Stress Conditional Fire Consequence",
        }
    )
    counts = (
        table.groupby("Facility Class", observed=True)
        .size()
        .rename("Selected Facilities")
        .reset_index()
    )
    diagnostics = {**admin_diagnostics, **facility_diagnostics}
    return table, counts, diagnostics


def build_definitions(diagnostics: dict[str, float]) -> pd.DataFrame:
    """Document the compact selection and interpretation rules."""
    rows = [
        (
            "Selection",
            "Four highest-priority facilities from each of five classes, producing a compact 20-row main-text table with balanced class representation.",
        ),
        (
            "Facility names",
            "Source names are omitted because English translations are not verified. Stable facility IDs remain internal matching keys but are omitted from the reader-facing table; source classification codes are retained where relevant.",
        ),
        (
            "Facility priority score",
            "Maximum of the prefecture-wide Combined Stress Consequence Rank and Accessibility Weakness Rank. This OR-style rule retains facilities with either high surrounding consequence or weak access.",
        ),
        (
            "Capacity or designation",
            "Type-specific source capacity or classification codes. NA means unavailable, not zero. Code meanings require confirmation against the source metadata before operational use.",
        ),
        (
            "Combined-stress consequence",
            "Relative cell-level screening score; not fire probability, burned area, or monetary loss.",
        ),
        (
            "Disrupted response time",
            "Nominal network travel time under the declared event-specific road-removal rule, capped at 30 minutes; not an observed dispatch time.",
        ),
        (
            "Backup base count",
            "Additional eligible candidate dispatch bases reachable within 10 minutes under the event-specific road rule.",
        ),
        (
            "Cell assignment",
            f"{int(diagnostics['assigned_facilities']):,} of {int(diagnostics['source_facilities']):,} source facilities were assigned to modeled cells; {int(diagnostics['fallback_facilities']):,} used a nearest-cell fallback no farther than {NEAREST_CELL_LIMIT_M:.0f} m.",
        ),
    ]
    return pd.DataFrame(rows, columns=["Item", "Definition or Boundary"])


def format_workbook(path: Path) -> None:
    """Apply compact formatting to the 20 x 9 main table."""
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="263746")
    band_fill = PatternFill("solid", fgColor="F1F4F6")
    sheet = workbook["Priority Facilities"]
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    widths = [12, 24, 27, 42, 17, 23, 19, 16, 18]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row_number in range(2, sheet.max_row + 1):
        if row_number % 2 == 0:
            for cell in sheet[row_number]:
                cell.fill = band_fill
        for column, cell in enumerate(sheet[row_number], start=1):
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="center",
                horizontal="left" if column in {2, 3, 4} else "center",
            )
        sheet.cell(row_number, 6).number_format = "0.000"
        sheet.cell(row_number, 7).number_format = "0.0"
        sheet.cell(row_number, 8).number_format = "0"
        sheet.cell(row_number, 9).number_format = "0.000"
        sheet.row_dimensions[row_number].height = 38
    sheet.row_dimensions[1].height = 62
    sheet.freeze_panes = "F2"
    sheet.auto_filter.ref = sheet.dimensions

    counts = workbook["Class Balance"]
    for cell in counts[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    counts.column_dimensions["A"].width = 31
    counts.column_dimensions["B"].width = 20
    counts.freeze_panes = "A2"

    definitions = workbook["Definitions"]
    for cell in definitions[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    definitions.column_dimensions["A"].width = 30
    definitions.column_dimensions["B"].width = 112
    for row_number in range(2, definitions.max_row + 1):
        definitions.cell(row_number, 1).font = Font(bold=True, color="263746")
        definitions.cell(row_number, 1).alignment = Alignment(vertical="top")
        definitions.cell(row_number, 2).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
        definitions.row_dimensions[row_number].height = 42
    definitions.freeze_panes = "A2"
    workbook.save(path)


def validate_english_only(path: Path) -> None:
    """Fail if any reader-facing workbook cell contains a Han character."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    hits = [
        (sheet.title, cell.coordinate, cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and HAN_PATTERN.search(cell.value)
    ]
    if hits:
        raise ValueError(f"Han characters remain in workbook: {hits[:5]}")


def main() -> None:
    table, counts, diagnostics = build_table()
    if table.shape != (OUTPUT_ROWS, 9):
        raise ValueError(f"Expected a {OUTPUT_ROWS} x 9 table, received {table.shape}")
    if not (counts["Selected Facilities"] == PER_CLASS_COUNT).all():
        raise ValueError("Each facility class must contribute exactly four rows")
    definitions = build_definitions(diagnostics)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        table.to_excel(writer, sheet_name="Priority Facilities", index=False)
        counts.to_excel(writer, sheet_name="Class Balance", index=False)
        definitions.to_excel(writer, sheet_name="Definitions", index=False)
    format_workbook(OUTPUT)
    validate_english_only(OUTPUT)
    print(f"Saved: {OUTPUT.relative_to(ROOT)}")
    print(
        "Diagnostics: "
        f"rows={len(table)}; columns={len(table.columns)}; "
        f"classes={table['Facility Class'].nunique()}; "
        f"assigned={int(diagnostics['assigned_facilities']):,}/"
        f"{int(diagnostics['source_facilities']):,}; "
        f"fallback={int(diagnostics['fallback_facilities']):,}; "
        "han_character_cells=0"
    )


if __name__ == "__main__":
    main()
