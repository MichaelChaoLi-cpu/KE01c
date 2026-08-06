#!/usr/bin/env python3
"""Generate a compact English-only top-20 fire-base value ranking."""

from __future__ import annotations

from pathlib import Path
import re

import numpy as np
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

from figure_fire_base_marginal_value_and_robustness import ENGLISH_STATION_LABELS
from table_municipality_fire_consequence_and_accessibility_summary import (
    ENGLISH_MUNICIPALITY,
)


ROOT = Path(__file__).resolve().parents[2]
VALUES_PATH = ROOT / "data/results/derived/fire_base_values.parquet"
CONVERGENCE_PATH = ROOT / "data/results/derived/fire_base_value_convergence.parquet"
OUTPUT = ROOT / "data/results/tables/Table_fire_base_value_ranking.xlsx"

OBJECTIVE = "Population"
TOP_BASE_COUNT = 20
HAN_PATTERN = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff]")

STATION_LABELS = {
    **ENGLISH_STATION_LABELS,
    "宇城広域連合北消防署": "Uki Regional North Fire Station",
    "熊本市消防局中央消防署北部出張所": "Kumamoto Central FS — North branch",
}


def build_table() -> tuple[pd.DataFrame, dict[str, float]]:
    """Select the population-objective top 20 under conservative robust value."""
    values = pd.read_parquet(VALUES_PATH)
    population = values.loc[
        values["Exposure Objective"].eq(OBJECTIVE)
        & values["Candidate Dispatch Base"].fillna(False)
    ].copy()
    expected_scenarios = {"normal", "central"}
    if set(population["Road Scenario"].unique()) != expected_scenarios:
        raise ValueError("Expected normal and central road scenarios")

    identity = (
        population.sort_values("Road Scenario", kind="stable")
        .drop_duplicates("Fire Base Name")
        [[
            "Fire Base Name",
            "Fire Base Type",
            "Municipality Code",
            "Robust Fire Base Value",
            "Robust Fire Base Rank",
            "Scenario Shapley Share IQR",
        ]]
        .copy()
    )
    shapley = population.pivot(
        index="Fire Base Name",
        columns="Road Scenario",
        values="Scenario Shapley Value Share",
    )
    central = (
        population.loc[population["Road Scenario"].eq("central")]
        .set_index("Fire Base Name")
        [["Leave-One-Out Value Share"]]
    )
    identity = identity.set_index("Fire Base Name").join(shapley).join(central)
    identity = identity.sort_values(
        ["Robust Fire Base Value", "Fire Base Name"],
        ascending=[False, True],
        kind="stable",
    ).head(TOP_BASE_COUNT)
    missing_labels = identity.index.difference(STATION_LABELS)
    if len(missing_labels):
        raise ValueError(f"Missing English station labels: {missing_labels.tolist()}")
    identity["Fire Base"] = identity.index.map(STATION_LABELS)
    identity["Municipality"] = identity["Municipality Code"].astype("string").map(
        ENGLISH_MUNICIPALITY
    )
    if identity["Municipality"].isna().any():
        raise ValueError("Missing English municipality labels in top-20 ranking")

    table = pd.DataFrame(
        {
            "Robust Rank": np.arange(1, len(identity) + 1, dtype=int),
            "Fire Base": identity["Fire Base"],
            "Base Type": identity["Fire Base Type"],
            "Municipality": identity["Municipality"],
            "Event Leave-One-Out Value Share (%)": 100
            * identity["Leave-One-Out Value Share"],
            "Normal-Road Shapley Value Share (%)": 100 * identity["normal"],
            "Event-Road Shapley Value Share (%)": 100 * identity["central"],
            "Robust Value Share (%)": 100 * identity["Robust Fire Base Value"],
            "Road-Scenario IQR (percentage points)": 100
            * identity["Scenario Shapley Share IQR"],
        }
    ).reset_index(drop=True)

    convergence = pd.read_parquet(CONVERGENCE_PATH)
    final_batch = convergence.loc[
        convergence["Permutation Count"].eq(convergence["Permutation Count"].max())
        & convergence["Exposure Objective"].eq(OBJECTIVE)
    ]
    diagnostics = {
        "eligible_bases": float(identity.shape[0] + (population["Fire Base Name"].nunique() - TOP_BASE_COUNT)),
        "permutation_count": float(population["Permutation Count"].max()),
        "all_converged": float(population["Shapley Converged"].all()),
        "final_rank_correlation_min": float(
            final_batch["Rank Correlation with Previous Batch"].min()
        ),
        "final_top_ten_overlap_min": float(final_batch["Top Ten Overlap"].min()),
    }
    return table, diagnostics


def build_definitions(diagnostics: dict[str, float]) -> pd.DataFrame:
    """Document value definitions and the current convergence limitation."""
    rows = [
        (
            "Selection",
            f"The 20 largest Robust Value Share estimates among {int(diagnostics['eligible_bases'])} eligible candidate dispatch bases for the population exposure objective.",
        ),
        (
            "Event leave-one-out share",
            "Share of total accessibility loss caused by removing the base from the full coalition under the event-specific road rule. It measures irreplaceability in the complete current system.",
        ),
        (
            "Shapley value share",
            "Average marginal accessibility contribution across sampled base coalitions. This is a cooperative-game Shapley value, not a machine-learning SHAP explanation.",
        ),
        (
            "Robust value share",
            "Median Shapley share across normal and event-specific road scenarios minus one-half of the road-scenario interquartile range.",
        ),
        (
            "Road-scenario IQR",
            "Interquartile range of the normal-road and event-road Shapley value shares, expressed in percentage points; larger values indicate greater road-scenario sensitivity.",
        ),
        (
            "Convergence limitation",
            f"All estimates use {int(diagnostics['permutation_count'])} sampled permutations and did not meet the pre-declared precision threshold. Final-batch rank correlation was at least {diagnostics['final_rank_correlation_min']:.3f}, and top-ten overlap was at least {diagnostics['final_top_ten_overlap_min']:.1%}. Rankings are provisional scenario estimates.",
        ),
        (
            "Interpretation boundary",
            "Values describe nominal accessibility contribution under declared scenarios. They do not measure actual vehicles, staffing, dispatch decisions, station quality, or causal effects.",
        ),
        (
            "English labels",
            "Reader-facing station labels are verified project translations used consistently with the accepted fire-base figure; source names remain in the derived data.",
        ),
    ]
    return pd.DataFrame(rows, columns=["Item", "Definition or Boundary"])


def format_workbook(path: Path) -> None:
    """Apply compact formatting to the 20 x 9 ranking."""
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="263746")
    band_fill = PatternFill("solid", fgColor="F1F4F6")
    sheet = workbook["Top 20 Fire Bases"]
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    widths = [12, 38, 20, 28, 22, 22, 22, 18, 22]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row_number in range(2, sheet.max_row + 1):
        if row_number % 2 == 0:
            for cell in sheet[row_number]:
                cell.fill = band_fill
        for column, cell in enumerate(sheet[row_number], start=1):
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="center",
                horizontal="left" if column in {2, 3, 4} else "center",
            )
        for column in range(5, 10):
            sheet.cell(row_number, column).number_format = "0.000"
        sheet.row_dimensions[row_number].height = 34
    sheet.row_dimensions[1].height = 62
    sheet.freeze_panes = "E2"
    sheet.auto_filter.ref = sheet.dimensions

    definitions = workbook["Definitions"]
    for cell in definitions[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    definitions.column_dimensions["A"].width = 31
    definitions.column_dimensions["B"].width = 112
    for row_number in range(2, definitions.max_row + 1):
        definitions.cell(row_number, 1).font = Font(bold=True, color="263746")
        definitions.cell(row_number, 1).alignment = Alignment(vertical="top")
        definitions.cell(row_number, 2).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
        definitions.row_dimensions[row_number].height = 44
    definitions.freeze_panes = "A2"
    workbook.save(path)


def validate_english_only(path: Path) -> None:
    """Fail if any reader-facing workbook cell contains a Han character."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    hits = [
        (sheet.title, cell.coordinate, cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and HAN_PATTERN.search(cell.value)
    ]
    if hits:
        raise ValueError(f"Han characters remain in workbook: {hits[:5]}")


def main() -> None:
    table, diagnostics = build_table()
    if table.shape != (TOP_BASE_COUNT, 9):
        raise ValueError(f"Expected a {TOP_BASE_COUNT} x 9 table, received {table.shape}")
    if not table["Robust Value Share (%)"].is_monotonic_decreasing:
        raise ValueError("Robust value ranking is not descending")
    definitions = build_definitions(diagnostics)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        table.to_excel(writer, sheet_name="Top 20 Fire Bases", index=False)
        definitions.to_excel(writer, sheet_name="Definitions", index=False)
    format_workbook(OUTPUT)
    validate_english_only(OUTPUT)
    print(f"Saved: {OUTPUT.relative_to(ROOT)}")
    print(
        "Diagnostics: "
        f"rows={len(table)}; columns={len(table.columns)}; "
        f"eligible_bases={int(diagnostics['eligible_bases'])}; "
        f"permutations={int(diagnostics['permutation_count'])}; "
        f"all_converged={bool(diagnostics['all_converged'])}; "
        "han_character_cells=0"
    )


if __name__ == "__main__":
    main()
