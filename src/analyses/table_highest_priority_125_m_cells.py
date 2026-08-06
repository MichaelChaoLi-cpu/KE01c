#!/usr/bin/env python3
"""Report the 50 highest-priority 125 m conditional-consequence cells."""

from __future__ import annotations

from pathlib import Path
import re

import geopandas as gpd
import numpy as np
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

from table_municipality_fire_consequence_and_accessibility_summary import (
    ENGLISH_MUNICIPALITY,
    assign_municipalities,
)


ROOT = Path(__file__).resolve().parents[2]
CONTEXT_PATH = ROOT / "data/results/derived/intervention_context_125m.parquet"
ADMIN_PATH = ROOT / "data/processed/administrative_areas_preprocessed.parquet"
LAND_USE_PATH = ROOT / "data/processed/mlit_land_use_zones_preprocessed.parquet"
OUTPUT = ROOT / "data/results/tables/Table_highest_priority_125_m_cells.xlsx"

PROJECTED_CRS = 6670
TOP_CELL_COUNT = 50
HAN_PATTERN = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff]")


def permitted_coverage_for_cells(
    cells: gpd.GeoDataFrame,
    land_use: gpd.GeoDataFrame,
) -> pd.DataFrame:
    """Return the permitted coverage ratio from the largest zone overlap."""
    zones = land_use[["Permitted Building Coverage Ratio", "Geometry"]].copy()
    zones["Permitted Building Coverage Ratio"] = pd.to_numeric(
        zones["Permitted Building Coverage Ratio"], errors="coerce"
    )
    zones = zones.loc[
        zones.geometry.notna() & zones["Permitted Building Coverage Ratio"].notna()
    ].copy()
    candidates = gpd.sjoin(
        cells[["Mesh Code", "Geometry"]],
        zones,
        how="left",
        predicate="intersects",
    ).reset_index(drop=True)
    candidates["Overlap Area (m2)"] = -1.0
    valid = candidates["index_right"].notna()
    zone_geometry = gpd.GeoSeries(
        candidates.loc[valid, "index_right"].map(zones.geometry),
        index=candidates.index[valid],
        crs=cells.crs,
    )
    candidates.loc[valid, "Overlap Area (m2)"] = (
        candidates.loc[valid].geometry.intersection(zone_geometry).area
    )
    return (
        candidates.sort_values(
            ["Mesh Code", "Overlap Area (m2)"],
            ascending=[True, False],
            kind="stable",
        )
        .drop_duplicates("Mesh Code")
        [["Mesh Code", "Permitted Building Coverage Ratio"]]
    )


def build_table() -> tuple[pd.DataFrame, dict[str, float]]:
    """Select, enrich, and format the top 50 combined-stress cells."""
    context = gpd.read_parquet(CONTEXT_PATH).to_crs(PROJECTED_CRS)
    administrative = gpd.read_parquet(ADMIN_PATH).to_crs(PROJECTED_CRS)
    land_use = gpd.read_parquet(LAND_USE_PATH).to_crs(PROJECTED_CRS)
    context["Mesh Code"] = context["Mesh Code"].astype("string")
    administrative["Municipality Code"] = administrative["Municipality Code"].astype(
        "string"
    )

    priority = (
        context.sort_values(
            ["Combined Stress Conditional Consequence", "Mesh Code"],
            ascending=[False, True],
            kind="stable",
        )
        .head(TOP_CELL_COUNT)
        .copy()
    )
    priority["Priority Rank"] = np.arange(1, len(priority) + 1, dtype=int)
    assignment, assignment_diagnostics = assign_municipalities(
        priority,
        administrative,
    )
    permitted = permitted_coverage_for_cells(priority, land_use)
    priority = priority.merge(
        assignment,
        on="Mesh Code",
        how="left",
        validate="one_to_one",
    ).merge(
        permitted,
        on="Mesh Code",
        how="left",
        validate="one_to_one",
    )
    priority["Municipality"] = priority["Municipality Code"].map(
        ENGLISH_MUNICIPALITY
    )
    if priority["Municipality"].isna().any():
        missing = priority.loc[
            priority["Municipality"].isna(), "Municipality Code"
        ].tolist()
        raise ValueError(f"Missing English municipality names for codes: {missing}")

    table = pd.DataFrame(
        {
            "Priority Rank": priority["Priority Rank"],
            "Mesh Code": priority["Mesh Code"],
            "Municipality Code": priority["Municipality Code"],
            "Municipality": priority["Municipality"],
            "Observed Building Footprint Coverage (%)": 100
            * priority["Observed Building Footprint Coverage Ratio"],
            "Permitted Building Coverage (%)": priority[
                "Permitted Building Coverage Ratio"
            ],
            "Mean Building Separation (m)": priority["Mean Building Separation (m)"],
            "Built Continuity (%)": 100 * priority["Built Continuity Index"],
            "Firebreak Share (%)": 100 * priority["Firebreak Share"],
            "Total Population": priority["Total Population"],
            "Disrupted Response Time (min)": priority[
                "Disrupted Response Time (min)"
            ],
            "Backup Fire Base Count": priority["Backup Fire Base Count"],
            "Water Constraint Scenario": priority["Water Constraint Scenario"],
            "Combined-Stress Conditional Fire Consequence": priority[
                "Combined Stress Conditional Consequence"
            ],
        }
    )
    diagnostics = {
        **assignment_diagnostics,
        "missing_permitted_coverage": float(
            table["Permitted Building Coverage (%)"].isna().sum()
        ),
        "bounded_water_cells": float(
            table["Water Constraint Scenario"]
            .astype("string")
            .str.contains("Bounded", case=False, na=False)
            .sum()
        ),
    }
    return table, diagnostics


def build_definitions(diagnostics: dict[str, float]) -> pd.DataFrame:
    """Document selection and interpretation rules in English."""
    rows = [
        (
            "Selection",
            "The 50 cells with the largest Combined-Stress Conditional Fire Consequence scores across Kumamoto Prefecture.",
        ),
        (
            "Priority rank",
            "Rank 1 has the largest relative combined-stress consequence score.",
        ),
        (
            "Observed building coverage",
            "Building-footprint area divided by 125 m cell area, expressed as a percentage.",
        ),
        (
            "Permitted building coverage",
            "Planning limit from the land-use zone with the largest overlap. NA means the cell is outside mapped zoning coverage; it is not zero.",
        ),
        (
            "Built continuity",
            "Relative within-cell continuity of mapped building footprints, expressed as a percentage; it does not identify building material or damage.",
        ),
        (
            "Firebreak share",
            "Mapped open-space or firebreak share of the cell, expressed as a percentage.",
        ),
        (
            "Disrupted response time",
            "Minimum nominal network travel time under the declared event-specific road-removal rule, capped at 30 minutes; not an observed dispatch time.",
        ),
        (
            "Backup base count",
            "Additional eligible candidate dispatch bases reachable within 10 minutes under the event-specific road rule.",
        ),
        (
            "Water constraint scenario",
            "Normal-reliance boundary or bounded non-reliance stress based on the mapped disruption area; not individual hydrant status.",
        ),
        (
            "Combined-stress consequence",
            "Relative screening score combining susceptibility, population exposure, event-specific accessibility, and the bounded water constraint. It is not a fire probability, burned area, or monetary loss.",
        ),
        (
            "Spatial assignment",
            f"{int(diagnostics['boundary_cells'])} selected cells required boundary review and {int(diagnostics['fallback_cells'])} required nearest-polygon assignment.",
        ),
    ]
    return pd.DataFrame(rows, columns=["Item", "Definition or Boundary"])


def format_workbook(path: Path) -> None:
    """Apply compact English-only workbook formatting."""
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="263746")
    band_fill = PatternFill("solid", fgColor="F1F4F6")
    sheet = workbook["Priority Cells"]
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    widths = [12, 17, 16, 27, 20, 19, 18, 16, 15, 15, 18, 16, 25, 24]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row_number in range(2, sheet.max_row + 1):
        if row_number % 2 == 0:
            for cell in sheet[row_number]:
                cell.fill = band_fill
        for column, cell in enumerate(sheet[row_number], start=1):
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="center",
                horizontal="left" if column in {4, 13} else "center",
            )
        for column in range(5, 10):
            sheet.cell(row_number, column).number_format = "0.0"
        if sheet.cell(row_number, 6).value is None:
            sheet.cell(row_number, 6).value = "NA"
            sheet.cell(row_number, 6).number_format = "General"
        sheet.cell(row_number, 10).number_format = "#,##0"
        sheet.cell(row_number, 11).number_format = "0.0"
        sheet.cell(row_number, 12).number_format = "0"
        sheet.cell(row_number, 14).number_format = "0.000"
        sheet.row_dimensions[row_number].height = 32
    sheet.row_dimensions[1].height = 62
    sheet.freeze_panes = "E2"
    sheet.auto_filter.ref = sheet.dimensions

    definitions = workbook["Definitions"]
    for cell in definitions[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    definitions.column_dimensions["A"].width = 31
    definitions.column_dimensions["B"].width = 112
    for row_number in range(2, definitions.max_row + 1):
        definitions.cell(row_number, 1).font = Font(bold=True, color="263746")
        definitions.cell(row_number, 1).alignment = Alignment(vertical="top")
        definitions.cell(row_number, 2).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
        definitions.row_dimensions[row_number].height = 40
    definitions.freeze_panes = "A2"
    definitions.auto_filter.ref = definitions.dimensions
    workbook.save(path)


def validate_english_only(path: Path) -> None:
    """Fail if any reader-facing workbook cell contains a Han character."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    hits = [
        (sheet.title, cell.coordinate, cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and HAN_PATTERN.search(cell.value)
    ]
    if hits:
        raise ValueError(f"Han characters remain in workbook: {hits[:5]}")


def main() -> None:
    table, diagnostics = build_table()
    if table.shape != (TOP_CELL_COUNT, 14):
        raise ValueError(f"Expected a {TOP_CELL_COUNT} x 14 table, received {table.shape}")
    if not table["Priority Rank"].equals(pd.Series(range(1, TOP_CELL_COUNT + 1))):
        raise ValueError("Priority ranks must be consecutive from 1 to 50")
    definitions = build_definitions(diagnostics)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        table.to_excel(writer, sheet_name="Priority Cells", index=False)
        definitions.to_excel(writer, sheet_name="Definitions", index=False)
    format_workbook(OUTPUT)
    validate_english_only(OUTPUT)
    print(f"Saved: {OUTPUT.relative_to(ROOT)}")
    print(
        "Diagnostics: "
        f"rows={len(table)}; columns={len(table.columns)}; "
        f"municipalities={table['Municipality'].nunique()}; "
        f"bounded_water_cells={int(diagnostics['bounded_water_cells'])}; "
        f"missing_permitted_coverage={int(diagnostics['missing_permitted_coverage'])}; "
        "han_character_cells=0"
    )


if __name__ == "__main__":
    main()
