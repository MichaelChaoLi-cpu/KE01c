#!/usr/bin/env python3
"""Generate a compact English-only robustness and sensitivity summary."""

from __future__ import annotations

from pathlib import Path
import re

import numpy as np
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

from road_failure_overlap_diagnostics import build_overlap_diagnostics


ROOT = Path(__file__).resolve().parents[2]
ROBUSTNESS = ROOT / "data/results/derived/robustness"
CONVERGENCE_PATH = (
    ROOT / "data/results/derived/road_reliability/road_reliability_convergence.parquet"
)
FORMAL_RELIABILITY_PATH = (
    ROOT / "data/results/derived/road_reliability/fire_service_reliability_125m.parquet"
)
FORMAL_REPLICATE_PATH = (
    ROOT / "data/results/derived/road_reliability/road_failure_replicate_metrics.parquet"
)
SUSCEPTIBILITY_PATH = ROOT / "data/results/derived/fire_susceptibility_125m.parquet"
CONSEQUENCE_PATH = ROOT / "data/results/derived/fire_consequence_125m.parquet"
OUTPUT = ROOT / "data/results/tables/Table_robustness_and_sensitivity_summary.xlsx"
HAN_PATTERN = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff]")

LD = "Length-dependent independent"
SC = "Spatially clustered"
HW = "Hazard-weighted"


def reliability_assessment(loss_pp: float, reference: bool = False) -> str:
    if reference:
        return "Reference"
    if loss_pp <= 1.0:
        return "Stable"
    if loss_pp <= 3.0:
        return "Moderately sensitive"
    return "Sensitive"


def base_assessment(rank_correlation: float, overlap: float) -> str:
    if rank_correlation >= 0.98 and overlap >= 0.80:
        return "Stable"
    if rank_correlation >= 0.95 and overlap >= 0.70:
        return "Moderately sensitive"
    return "Sensitive"


def intervention_assessment(lower_quartile: float, worst_state: float) -> str:
    if lower_quartile >= 0.95 and worst_state >= 0.90:
        return "Stable"
    if lower_quartile >= 0.85 and worst_state >= 0.70:
        return "Moderately sensitive"
    return "Sensitive"


def percentile_rank(values: pd.Series) -> np.ndarray:
    """Reproduce the empirical percentile transform used by the main analysis."""
    series = pd.Series(values, dtype="float64")
    valid_n = int(series.notna().sum())
    if valid_n <= 1:
        return np.full(len(series), np.nan, dtype=float)
    return ((series.rank(method="average", na_option="keep") - 1) / (valid_n - 1)).to_numpy()


def rank_stability(reference: np.ndarray, alternative: np.ndarray) -> tuple[float, float]:
    """Return rank correlation and top-decile Jaccard membership overlap."""
    valid = np.isfinite(reference) & np.isfinite(alternative)
    reference = np.asarray(reference)[valid]
    alternative = np.asarray(alternative)[valid]
    if len(reference) < 2:
        raise ValueError("Too few valid observations for a rank-stability check")
    reference_rank = pd.Series(reference).rank(method="average").to_numpy()
    alternative_rank = pd.Series(alternative).rank(method="average").to_numpy()
    correlation = float(np.corrcoef(reference_rank, alternative_rank)[0, 1])
    top_n = max(1, int(np.ceil(0.10 * len(reference))))
    reference_top = set(np.argpartition(reference, -top_n)[-top_n:].tolist())
    alternative_top = set(np.argpartition(alternative, -top_n)[-top_n:].tolist())
    overlap = len(reference_top & alternative_top) / len(reference_top | alternative_top)
    return correlation, float(overlap)


def build_table() -> pd.DataFrame:
    """Combine available robustness results and explicit evidence gaps in 20 rows."""
    road = pd.read_parquet(ROBUSTNESS / "road_mechanism_summary.parquet")
    base = pd.read_parquet(
        ROBUSTNESS / "fire_base_leave_one_out_state_stability.parquet"
    )
    intervention = pd.read_parquet(
        ROBUSTNESS / "road_intervention_priority_stability.parquet"
    )
    convergence = pd.read_parquet(CONVERGENCE_PATH)
    formal_reliability = pd.read_parquet(FORMAL_RELIABILITY_PATH)
    formal_replicates = pd.read_parquet(FORMAL_REPLICATE_PATH)
    susceptibility = pd.read_parquet(SUSCEPTIBILITY_PATH)
    consequence = pd.read_parquet(CONSEQUENCE_PATH)

    formal_checkpoint = convergence.loc[
        convergence["Simulation Replicates"].eq(1000)
        & convergence["Road Failure Model"].astype("string").str.casefold().eq(LD.casefold())
        & convergence["Expected Failed Road Length Share"].eq(0.03)
    ]
    if len(formal_checkpoint) != 1:
        raise ValueError("Expected one formal 1,000-replicate 3% checkpoint")
    if (
        len(formal_replicates) != 1000
        or formal_replicates["Simulation Replicate"].nunique() != 1000
        or not formal_replicates["Road Failure Model"]
        .astype("string")
        .str.casefold()
        .eq(LD.casefold())
        .all()
        or not formal_replicates["Expected Failed Road Length Share"].eq(0.03).all()
    ):
        raise ValueError("Formal road-reliability replicate cache is not the declared 1,000-state 3% experiment")
    if (
        not formal_reliability["Road Failure Model"]
        .astype("string")
        .str.casefold()
        .eq(LD.casefold())
        .all()
        or not formal_reliability["Expected Failed Road Length Share"].eq(0.03).all()
        or not formal_reliability["Simulation Replicates"].eq(1000).all()
    ):
        raise ValueError("Formal grid-reliability cache does not match the declared 1,000-state 3% experiment")

    formal_item = formal_checkpoint.iloc[0]
    formal_p90 = formal_reliability["P90 Response Time (min)"].to_numpy(dtype=float)
    finite_formal_p90 = formal_p90[np.isfinite(formal_p90)]
    if finite_formal_p90.size == 0:
        raise ValueError("Formal grid-reliability cache has no finite supported P90 values")
    formal_timely = float(
        formal_item["Population-Weighted Timely Response Probability"]
    )
    formal_median_p90 = float(np.median(finite_formal_p90))
    formal_disconnection = float(
        formal_reliability["Disconnection Probability"].mean()
    )
    formal_p90_supported = float(np.isfinite(formal_p90).mean())

    rows: list[dict[str, str]] = []
    reliability_specs = [
        (LD, 0.01, "Primary severity path"),
        (LD, 0.03, "Primary severity path"),
        (LD, 0.05, "Primary severity path"),
        (LD, 0.10, "Primary stress boundary"),
    ]
    one_percent = road.loc[
        road["Expected Failed Road Length Share"].eq(0.01)
    ].set_index("Road Failure Model")["Mean Timely Response Probability"]
    for model, share, specification in reliability_specs:
        match = road.loc[
            road["Road Failure Model"].eq(model)
            & road["Expected Failed Road Length Share"].eq(share)
        ]
        if len(match) != 1:
            raise ValueError(f"Missing road robustness row for {model}, {share}")
        item = match.iloc[0]
        is_formal = model == LD and share == 0.03
        timely = formal_timely if is_formal else float(
            item["Mean Timely Response Probability"]
        )
        loss_pp = 100 * (float(one_percent.loc[model]) - timely)
        if is_formal:
            secondary_result = (
                f"Median supported-cell P90 {formal_median_p90:.2f} min"
            )
            interpretation = (
                "Formal 1,000-replicate estimate; "
                f"timely-response loss versus the 100-replicate 1% sensitivity reference: {loss_pp:.2f} percentage points; "
                f"mean cell disconnection probability: {100 * formal_disconnection:.2f}%; "
                f"cells with a finite supported P90: {100 * formal_p90_supported:.2f}%."
            )
        else:
            secondary_result = (
                f"Mean connected-demand P90 {item['Mean P90 Response Time (min)']:.2f} min"
            )
            interpretation = (
                "100-replicate sensitivity estimate; "
                f"timely-response loss versus the same-mechanism 1% case: {loss_pp:.2f} percentage points; "
                f"mean disconnected demand share: {100 * item['Mean Disconnected Demand Share']:.2f}%."
            )
        rows.append(
            {
                "Domain": "Road-response reliability",
                "Specification": specification,
                "Failure Model": model,
                "Failed Road Length": f"{100 * share:.0f}%",
                "Primary Result": f"Mean timely response {100 * timely:.2f}%",
                "Secondary Result": secondary_result,
                "Assessment": reliability_assessment(loss_pp, reference=share == 0.01),
                "Interpretation": interpretation,
            }
        )

    alternative = road.loc[
        road["Road Failure Model"].isin([SC, HW])
        & road["Expected Failed Road Length Share"].eq(0.03)
    ].copy()
    alternative["Loss pp"] = alternative.apply(
        lambda item: 100
        * (
            float(one_percent.loc[item["Road Failure Model"]])
            - float(item["Mean Timely Response Probability"])
        ),
        axis=1,
    )
    worst = alternative.loc[alternative["Loss pp"].idxmax()]
    rows.append(
        {
            "Domain": "Road-response reliability",
            "Specification": "Worst alternative mechanism",
            "Failure Model": "SC / HW",
            "Failed Road Length": "3%",
            "Primary Result": f"Minimum timely response {100 * alternative['Mean Timely Response Probability'].min():.2f}%",
            "Secondary Result": f"Maximum mean P90 {alternative['Mean P90 Response Time (min)'].max():.2f} min",
            "Assessment": reliability_assessment(float(worst["Loss pp"])),
            "Interpretation": (
                f"Largest same-mechanism loss versus 1% is {worst['Loss pp']:.2f} percentage points "
                f"under {worst['Road Failure Model']}."
            ),
        }
    )

    overlap_diagnostics = build_overlap_diagnostics((0.03,))
    if len(overlap_diagnostics) != 2:
        raise ValueError("Expected two nominal-input overlap diagnostic rows")
    for item in overlap_diagnostics.itertuples(index=False):
        rows.append(
            {
                "Domain": "Road-failure overlap diagnostic",
                "Specification": "Nominal input versus effective addition",
                "Failure Model": item[0],
                "Failed Road Length": "3%",
                "Primary Result": (
                    f"Expected overlap {100 * item[3]:.2f}%"
                ),
                "Secondary Result": (
                    f"Added {100 * item[4]:.2f}% of pre-mask; "
                    f"{100 * item[5]:.2f}% of event-available"
                ),
                "Assessment": "Interpretation boundary",
                "Interpretation": (
                    "Analytical diagnostic only; retained routing states are not "
                    "an equal-net-damage mechanism comparison."
                ),
            }
        )

    for model in [LD, SC, HW]:
        subset = base.loc[base["Road Failure Model"].eq(model)]
        if subset["Exposure Objective"].nunique() != 3 or len(subset) != 30:
            raise ValueError(f"Expected 10 states for each of three objectives for {model}")
        minimum_correlation = float(subset["Fire Base Rank Correlation"].min())
        minimum_overlap = float(subset["Top Ten Fire Base Overlap"].min())
        rows.append(
            {
                "Domain": "Fire-base priority",
                "Specification": "Worst of 10 states × 3 objectives",
                "Failure Model": model,
                "Failed Road Length": "3%",
                "Primary Result": f"Minimum rank correlation {minimum_correlation:.3f}",
                "Secondary Result": f"Minimum top-10 retained {100 * minimum_overlap:.0f}%",
                "Assessment": base_assessment(minimum_correlation, minimum_overlap),
                "Interpretation": (
                    f"Median rank correlation {subset['Fire Base Rank Correlation'].median():.3f}; "
                    f"median top-10 overlap {100 * subset['Top Ten Fire Base Overlap'].median():.0f}%."
                ),
            }
        )

    for model in [LD, SC, HW]:
        subset = intervention.loc[intervention["Road Failure Model"].eq(model)]
        counts = subset.groupby("Road Priority Rule")["Simulation Replicate"].nunique()
        if set(counts.index) != {"Section count", "Length-aware"} or not counts.eq(100).all():
            raise ValueError(f"Expected 100 fixed-plan states for each road-priority rule under {model}")
        by_rule = subset.groupby("Road Priority Rule")[
            "Retained Protection Gain Share"
        ].agg(
            Median="median",
            **{
                "Lower Quartile": lambda values: values.quantile(0.25),
                "Worst State": "min",
            },
        )
        limiting_rule = str(by_rule["Lower Quartile"].idxmin())
        median_retained = float(by_rule["Median"].min())
        lower_quartile = float(by_rule["Lower Quartile"].min())
        worst_state = float(by_rule["Worst State"].min())
        rows.append(
            {
                "Domain": "Fixed road-intervention plan",
                "Specification": "100 states per priority rule",
                "Failure Model": model,
                "Failed Road Length": "3%",
                "Primary Result": f"Lowest median retained gain {100 * median_retained:.2f}%",
                "Secondary Result": (
                    f"Lowest Q25 {100 * lower_quartile:.2f}%; "
                    f"worst state {100 * worst_state:.2f}%"
                ),
                "Assessment": intervention_assessment(lower_quartile, worst_state),
                "Interpretation": (
                    f"The limiting lower-quartile result is for {limiting_rule}; "
                    "the same event-road bundle is evaluated in every state without state-specific re-optimization."
                ),
            }
        )

    checkpoint = formal_item
    mean_change = float(checkpoint["Mean Absolute Grid Probability Change"])
    jaccard = float(checkpoint["Top-Decile Priority Jaccard"])
    assessment = "Stable" if mean_change <= 0.001 and jaccard >= 0.90 else "Sensitive"
    rows.append(
        {
            "Domain": "Road-reliability convergence",
            "Specification": "1,000-replicate checkpoint",
            "Failure Model": LD,
            "Failed Road Length": "3%",
            "Primary Result": f"Mean probability change {mean_change:.4f}",
            "Secondary Result": f"Top-decile Jaccard {jaccard:.3f}",
            "Assessment": assessment,
            "Interpretation": (
                f"Change relative to the {int(checkpoint['Previous Checkpoint'])}-replicate checkpoint; "
                f"maximum cell probability change: {checkpoint['Maximum Absolute Grid Probability Change']:.4f}."
            ),
        }
    )

    coverage_rank = percentile_rank(
        susceptibility["Observed Building Footprint Coverage Ratio"]
    )
    continuity_rank = percentile_rank(susceptibility["Built Continuity Index"])
    separation_rank = percentile_rank(susceptibility["Mean Building Separation (m)"])
    firebreak_rank = percentile_rank(susceptibility["Firebreak Share"])
    inverse_separation = 1 - separation_rank
    inverse_separation[~np.isfinite(inverse_separation)] = 0
    continuity_rank[susceptibility["Building Count"].to_numpy() < 2] = 0
    components = {
        "Building coverage": coverage_rank,
        "Built continuity": continuity_rank,
        "Inverse separation": inverse_separation,
        "Limited firebreak": 1 - firebreak_rank,
    }
    reference_susceptibility = susceptibility[
        "Conditional Spread Susceptibility"
    ].to_numpy(dtype=float)
    reconstructed = np.nanmean(np.column_stack(list(components.values())), axis=1)
    if not np.allclose(reference_susceptibility, reconstructed, equal_nan=True, atol=1e-10):
        raise ValueError("Stored susceptibility does not match the four-component definition")
    for omitted, _ in components.items():
        alternative_score = np.nanmean(
            np.column_stack([value for name, value in components.items() if name != omitted]),
            axis=1,
        )
        correlation, overlap = rank_stability(reference_susceptibility, alternative_score)
        rows.append(
            {
                "Domain": "Built-form susceptibility",
                "Specification": f"Omit {omitted.lower()}",
                "Failure Model": "Component ablation",
                "Failed Road Length": "—",
                "Primary Result": f"Rank correlation {correlation:.3f}",
                "Secondary Result": f"Top-decile Jaccard {overlap:.3f}",
                "Assessment": "—",
                "Interpretation": "Equal-weight reference compared with the same score after removing one component.",
            }
        )

    reference_consequence = consequence[
        "Population Conditional Fire Consequence"
    ].to_numpy(dtype=float)
    susceptibility_score = consequence["Conditional Spread Susceptibility"].to_numpy(dtype=float)
    population_exposure = consequence["Population Exposure"].to_numpy(dtype=float)
    access_penalty = consequence["Accessibility Penalty"].to_numpy(dtype=float)
    rebuilt_consequence = susceptibility_score * population_exposure * (1 + access_penalty)
    if not np.allclose(reference_consequence, rebuilt_consequence, equal_nan=True, atol=1e-10):
        raise ValueError("Stored population consequence does not match the declared reference")
    for multiplier in [0.5, 1.5]:
        alternative_consequence = (
            susceptibility_score * population_exposure * (1 + multiplier * access_penalty)
        )
        correlation, overlap = rank_stability(reference_consequence, alternative_consequence)
        rows.append(
            {
                "Domain": "Conditional consequence",
                "Specification": f"Accessibility multiplier = {multiplier:.1f}",
                "Failure Model": "Parameter sensitivity",
                "Failed Road Length": "—",
                "Primary Result": f"Rank correlation {correlation:.3f}",
                "Secondary Result": f"Top-decile Jaccard {overlap:.3f}",
                "Assessment": "—",
                "Interpretation": "Population consequence ranking compared with the reference multiplier of 1.0.",
            }
        )

    table = pd.DataFrame(rows)
    table["Failed Road Length"] = table["Failed Road Length"].map(
        lambda value: f"{value} nominal" if str(value).endswith("%") else value
    )
    table = table.rename(columns={"Failed Road Length": "Nominal Road Input"})
    if table.shape != (20, 8):
        raise ValueError(f"Expected a 20 x 8 robustness table, received {table.shape}")
    if table.isna().any().any():
        raise ValueError("Robustness summary contains missing cells")
    return table


def build_definitions() -> pd.DataFrame:
    rows = [
        (
            "Scope",
            "The table reports exact values that support, but do not duplicate, the accepted Scenario and Parameter Robustness figure.",
        ),
        (
            "Road-response rows",
            "The percentages are nominal inputs calibrated over the eligible pre-mask network. The length-dependent nominal 3% row uses the formal 1,000-replicate grid-reliability and convergence outputs. Its timely-response result is population weighted, and its secondary result is the median finite cell-level unconditional P90 among cells satisfying the below-10% disconnection reporting rule. Other severity and mechanism rows use 100-replicate sensitivity runs and report a connected-demand within-state P90 diagnostic.",
        ),
        (
            "Overlap diagnostic rows",
            "Expected overlap and effective added unavailable shares are calculated analytically from existing section probabilities and event-exposed length without rerouting. They show why mechanisms at the same nominal input are bounded stress scenarios rather than equal-net-damage comparisons.",
        ),
        (
            "Fire-base rows",
            "Deterministic one-base removal is evaluated across 10 predeclared road states under the nominal 3% input for each of three exposure objectives and each retained failure mechanism. Each row reports the worst result across 30 state-objective pairs.",
        ),
        (
            "Intervention rows",
            "The section-count and length-aware bundles selected under the event-specific road case are each evaluated unchanged across 100 predeclared road states per failure mechanism under the nominal 3% input. Rows report the lower result across the two rules and do not imply state-specific re-optimization.",
        ),
        (
            "Built-form rows",
            "Each component-ablation row removes one of the four equal-weight susceptibility components and compares the new ranking with the reference ranking. Exact statistics are reported without a categorical assessment threshold.",
        ),
        (
            "Consequence rows",
            "Population consequence is recalculated with accessibility-penalty multipliers of 0.5 and 1.5 and compared with the reference multiplier of 1.0. Exact statistics are reported without a categorical assessment threshold.",
        ),
        (
            "Road reliability assessment",
            "Stable means timely-response loss of at most 1 percentage point relative to the same-mechanism 1% case; moderately sensitive means more than 1 and at most 3 points; sensitive means more than 3 points.",
        ),
        (
            "Fire-base assessment",
            "Stable requires rank correlation at least 0.98 and top-10 overlap at least 80%; moderately sensitive requires correlation at least 0.95 and overlap at least 70%; otherwise sensitive.",
        ),
        (
            "Intervention assessment",
            "Stable requires a lower-quartile retained protection gain of at least 95% and a worst-state value of at least 90%; moderately sensitive requires at least 85% and 70%, respectively; otherwise sensitive.",
        ),
        (
            "Convergence assessment",
            "Stable at the final checkpoint requires mean absolute grid-probability change at most 0.001 and top-decile Jaccard at least 0.90.",
        ),
        (
            "Interpretation boundary",
            "Failure shares and mechanisms are declared nominal pre-mask scenario inputs, not observed damage rates, engineering-calibrated failure probabilities, or equal-net-additional-damage comparisons.",
        ),
    ]
    return pd.DataFrame(rows, columns=["Item", "Definition or Boundary"])


def format_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="263746")
    band_fill = PatternFill("solid", fgColor="F1F4F6")
    assessment_fill = {
        "Reference": PatternFill("solid", fgColor="E8EEF2"),
        "Stable": PatternFill("solid", fgColor="E5F2EA"),
        "Moderately sensitive": PatternFill("solid", fgColor="FFF1D6"),
        "Sensitive": PatternFill("solid", fgColor="F8DEDE"),
        "Limited evidence": PatternFill("solid", fgColor="FFF1D6"),
        "Interpretation boundary": PatternFill("solid", fgColor="E8EEF2"),
    }
    sheet = workbook["Robustness Summary"]
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    widths = [27, 36, 31, 18, 31, 31, 23, 62]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for row_number in range(2, sheet.max_row + 1):
        if row_number % 2 == 0:
            for cell in sheet[row_number]:
                cell.fill = band_fill
        for column, cell in enumerate(sheet[row_number], start=1):
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
                horizontal="center" if column in {4, 7} else "left",
            )
        assessment = sheet.cell(row_number, 7)
        assessment.fill = assessment_fill.get(assessment.value, band_fill)
        assessment.font = Font(bold=True, color="263746")
        sheet.row_dimensions[row_number].height = 52
    sheet.row_dimensions[1].height = 48
    sheet.freeze_panes = "E2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_title_rows = "1:1"
    sheet.print_area = sheet.dimensions
    sheet.page_margins.left = 0.2
    sheet.page_margins.right = 0.2
    sheet.page_margins.top = 0.3
    sheet.page_margins.bottom = 0.3

    definitions = workbook["Definitions"]
    for cell in definitions[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    definitions.column_dimensions["A"].width = 32
    definitions.column_dimensions["B"].width = 112
    for row_number in range(2, definitions.max_row + 1):
        definitions.cell(row_number, 1).font = Font(bold=True, color="263746")
        definitions.cell(row_number, 1).alignment = Alignment(vertical="top")
        definitions.cell(row_number, 2).alignment = Alignment(wrap_text=True, vertical="top")
        definitions.row_dimensions[row_number].height = 46
    definitions.freeze_panes = "A2"
    definitions.sheet_properties.pageSetUpPr.fitToPage = True
    definitions.page_setup.orientation = "landscape"
    definitions.page_setup.paperSize = definitions.PAPERSIZE_A3
    definitions.page_setup.fitToWidth = 1
    definitions.page_setup.fitToHeight = 1
    definitions.print_title_rows = "1:1"
    definitions.print_area = definitions.dimensions
    definitions.page_margins.left = 0.2
    definitions.page_margins.right = 0.2
    definitions.page_margins.top = 0.3
    definitions.page_margins.bottom = 0.3
    workbook.save(path)


def validate_english_only(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    hits = [
        (sheet.title, cell.coordinate, cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and HAN_PATTERN.search(cell.value)
    ]
    if hits:
        raise ValueError(f"Han characters remain in workbook: {hits[:5]}")


def main() -> None:
    table = build_table()
    if table.shape != (20, 8):
        raise ValueError(f"Expected a 20 x 8 table, received {table.shape}")
    definitions = build_definitions()
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        table.to_excel(writer, sheet_name="Robustness Summary", index=False)
        definitions.to_excel(writer, sheet_name="Definitions", index=False)
    format_workbook(OUTPUT)
    validate_english_only(OUTPUT)
    print(f"Saved: {OUTPUT.relative_to(ROOT)}")
    print(
        "Diagnostics: "
        f"rows={len(table)}; columns={len(table.columns)}; "
        f"assessments={table['Assessment'].value_counts().to_dict()}; "
        "han_character_cells=0"
    )


if __name__ == "__main__":
    main()
